from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
try:
    from playwright_stealth import Stealth
except ImportError:
    Stealth = None
from collections import Counter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd
import random
import time
import shutil
import re
import os
import json
import requests
import sys
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode


# 페이지당 최대 크롤링 상품 수 (0이면 제한 없음)
MAX_PRODUCTS_PER_PAGE = int(os.getenv("MAX_PRODUCTS_PER_PAGE", "0") or 0)


# .env 로더 (python-dotenv 미설치 시 최소 파서)
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None


def fallback_load_dotenv(dotenv_path):
    path = Path(dotenv_path)
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        return False
    except OSError as exc:
        print(f".env 파일을 읽는 중 오류가 발생했습니다: {exc}")
        return False

    loaded = False
    saw_assignments = False
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        saw_assignments = True
        key = key.strip()
        value = value.strip().strip("\"'")
        if not key:
            continue
        if key not in os.environ:
            os.environ[key] = value
            loaded = True
    return loaded or saw_assignments


SCRIPT_DIR = Path(__file__).resolve().parent
LOG_FILE = SCRIPT_DIR / "log.txt"


def resolve_category_path() -> Path:
    # 우선 로컬 경로, 없으면 상위 디렉터리(AGENTS.md 가이드에 맞춤)
    local = SCRIPT_DIR / "naver_category.xlsx"
    parent = SCRIPT_DIR.parent / "naver_category.xlsx"
    if local.exists():
        return local
    if parent.exists():
        return parent
    # 마지막으로 로컬 경로를 반환(실패 시 런타임에서 에러 메시지 제공)
    return local


NAVER_CATEGORY_PATH = resolve_category_path()
CRAWLER_DRY_RUN = os.getenv("CRAWLER_DRY_RUN", "0").lower() in {"1", "true", "yes"}

# 선택 페이지만 크롤링하는 디버그용 옵션(예: CRAWL_ONLY_PAGES="51,59").
# 지정되지 않으면 기존 범위(global_start_page~global_last_page) 전체를 처리합니다.
_only_pages_raw = os.getenv("CRAWL_ONLY_PAGES", "").strip()
CRAWL_ONLY_PAGES = None
if _only_pages_raw:
    try:
        parts = re.split(r"[\s,;]+", _only_pages_raw)
        CRAWL_ONLY_PAGES = [int(p) for p in parts if p]
    except Exception as exc:
        print(f"CRAWL_ONLY_PAGES 파싱 실패({_only_pages_raw}): {exc}")
        CRAWL_ONLY_PAGES = None

# 페이지 번호 이동 시 URL 쿼리 파라미터로 강제 점프 시도 여부
# 기본값은 비활성화(네이버는 URL 파라미터만으로 DOM이 바뀌지 않는 경우가 많음)
PAGE_JUMP_BY_QUERY = False  # Forced OFF: Naver does not allow page navigation via URL params

# 페이지 이동 과정 스크린샷 저장(디버깅 용도)
# 기본값 비활성화: 요청에 따라 캡처 중단
PAGINATION_DEBUG_SHOTS = os.getenv("PAGINATION_DEBUG_SHOTS", "0").lower() in {"1", "true", "yes"}

# 페이지네이션 전략: auto(기본) | next_only('다음'만 반복)
PAGINATION_STRATEGY = os.getenv("PAGINATION_STRATEGY", "auto").lower().strip()

def debug_shot(page, label):
    if not PAGINATION_DEBUG_SHOTS:
        return
    try:
        ts = time.strftime("%Y%m%d_%H%M%S")
        dbg = (SCRIPT_DIR / "debug")
        dbg.mkdir(exist_ok=True)
        path = dbg / f"pagination_{ts}_{label}.png"
        page.screenshot(path=str(path), full_page=True)
        print(f"Saved screenshot: {path}")
    except Exception as exc:
        print(f"Failed to take screenshot({label}): {exc}")

STEALTH_HELPER = Stealth() if Stealth is not None else None
if STEALTH_HELPER is None:
    print(
        "playwright_stealth 모듈에서 Stealth 클래스를 불러오지 못했습니다. "
        "탐지 회피 스크립트가 적용되지 않으니 chromium 환경에서는 추가 점검이 필요합니다."
    )


if load_dotenv:
    load_dotenv()
else:
    _ = fallback_load_dotenv(SCRIPT_DIR / ".env")


def first_available(node, selectors):
    for selector in selectors:
        try:
            element = node.query_selector(selector)
        except PlaywrightTimeoutError:
            continue
        if element:
            return element
    return None


def find_elements(page, selectors):
    for selector in selectors:
        try:
            elements = page.query_selector_all(selector)
        except PlaywrightTimeoutError:
            continue
        if elements:
            print(f"Selector '{selector}' matched {len(elements)} elements.")
            return elements
    return []


def save_debug_snapshot(page, prefix):
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    path = (SCRIPT_DIR / "debug")
    path.mkdir(exist_ok=True)
    out = path / f"{prefix}_{timestamp}.html"
    try:
        out.write_text(page.content(), encoding="utf-8")
        print(f"Saved debug snapshot: {out}")
    except Exception as exc:
        print(f"Failed to save debug snapshot: {exc}")


def extract_price_from_text(raw_text):
    match = re.search(r"([\d,]+)\s*원", raw_text)
    if match:
        return match.group(1).replace("\u200b", "").strip()
    digits = re.sub(r"[^\d]", "", raw_text)
    if not digits:
        return "N/A"
    try:
        return "{:,}".format(int(digits))
    except ValueError:
        return "N/A"


def update_query_params(url, **params):
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({key: value for key, value in params.items() if value is not None})
    new_query = urlencode(query, doseq=True)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))


class Tee(object):
    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()

    def flush(self):
        for f in self.files:
            f.flush()


f = LOG_FILE.open('w', encoding='utf-8')
original = sys.stdout
sys.stdout = Tee(sys.stdout, f)


base_url = "https://smartstore.naver.com"


def product_list_crawl(context, df, read_excel_path, seen_urls):
    page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(page)

    raw_url = 'https://smartstore.naver.com/joypapa_/category/ALL?st=RECENT&dt=BIG_IMAGE&size=80'
    original_url = update_query_params(raw_url, page=None)
    page.goto(original_url)
    page.wait_for_load_state("load")
    page.wait_for_load_state("networkidle")

    global_start_page = 51
    global_last_page = 59
    # 디버그: 특정 페이지만 요청된 경우 범위를 해당 값으로 축소
    if CRAWL_ONLY_PAGES:
        try:
            global_start_page = min(CRAWL_ONLY_PAGES)
            global_last_page = max(CRAWL_ONLY_PAGES)
        except Exception:
            pass
    shopname = raw_url.split('/')[3]
    shopnumber = raw_url.split('/')[5].split('?')[0]

    home_dir = Path.home()
    output_folder = home_dir / 'Desktop' / 'excel_output'
    output_folder.mkdir(parents=True, exist_ok=True)

    pagination_button_labels = {
        "next": ["다음", "다음 페이지", "다음페이지", ">"],
        "prev": ["이전", "이전 페이지", "이전페이지", "<"]
    }

    # 검증 모드: 특정 페이지의 첫 상품을 열어 기대 URL/이름 확인
    verify_target_page = None
    verify_expected_url = (os.getenv("VERIFY_FIRST_PRODUCT_URL") or "").strip() or None
    verify_expected_name = (os.getenv("VERIFY_FIRST_PRODUCT_NAME") or "").strip() or None
    try:
        verify_target_page = int(os.getenv("VERIFY_TARGET_PAGE", "") or 0) or None
    except Exception:
        verify_target_page = None

    def scroll_to_pagination():
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        except PlaywrightTimeoutError:
            pass
        time.sleep(0.8)
        # 하단 스크롤 후 간단 스크린샷
        debug_shot(page, "scrolled_bottom")

    def wait_pagination_ready(timeout_ms=8000):
        # 페이지네이션 컨테이너 또는 숫자/버튼이 나타날 때까지 대기
        selectors = [
            "div[data-shp-area='list.pgn'][role='menubar']",
            "div[data-shp-contents-type='pgn'][role='menubar']",
            "div[data-shp-area-id='pgn'][role='menubar']",
        ]
        end = time.time() + (timeout_ms / 1000.0)
        while time.time() < end:
            container = find_pagination_container()
            if container:
                try:
                    if container.query_selector("a[role='menuitem'],a[role='button']"):
                        return True
                except Exception:
                    pass
            time.sleep(0.3)
        return False

    def get_first_list_href():
        try:
            el = page.query_selector("a[href*='/products/']")
            if el:
                href = el.get_attribute("href")
                return href
        except Exception:
            pass
        return None

    def find_pagination_container():
        candidates = [
            "div[data-shp-area='list.pgn'][role='menubar']",
            "div[data-shp-contents-type='pgn'][role='menubar']",
            "div[data-shp-area-id='pgn'][role='menubar']",
            # 폴백들
            "nav[aria-label*='페이지']",
            "nav[aria-label*='pagination']",
            "nav[role='navigation']",
            "div[class*='Pagination']",
            "div[class*='paginate']",
            "div[class*='paging']",
        ]
        for sel in candidates:
            try:
                elem = page.query_selector(sel)
            except Exception:
                elem = None
            if elem:
                return elem
        return None

    def get_pgn_from_container():
        """컨테이너의 data-shp-filter_con 속성에서 pgn 값을 파싱(네이버 구조 특화)."""
        try:
            container = find_pagination_container()
            if not container:
                return None
            raw = container.get_attribute("data-shp-filter_con")
            if not raw:
                return None
            # HTML 인코딩된 문자열 처리
            raw = raw.replace('&quot;', '"')
            data = json.loads(raw)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict) and item.get('key') in {'pgn', 'page', 'pageNum'}:
                        val = item.get('value')
                        try:
                            return int(re.findall(r"\d+", str(val))[0])
                        except Exception:
                            pass
        except Exception:
            return None
        return None

    def find_page_link_in_container(target_page):
        container = find_pagination_container()
        if container is None:
            return None

        # Strict container-only search to avoid clicking product links.
        for sel in ["a[role='menuitem']", "button[role='menuitem']"]:
            try:
                for node in container.query_selector_all(sel):
                    try:
                        t = (node.inner_text() or "").strip()
                    except Exception:
                        continue
                    if t.isdigit() and int(t) == target_page:
                        return node
            except Exception:
                pass

        for sel in ["a", "button", "span", "li"]:
            try:
                for node in container.query_selector_all(sel):
                    try:
                        t = (node.inner_text() or "").strip()
                    except Exception:
                        continue
                    if not t:
                        continue
                    if t.isdigit() and int(t) == target_page:
                        return node
            except Exception:
                pass

        return None

    def ensure_group_has_page(target_page, max_group_hops=20):
        """타겟 숫자 링크가 현재 보이는 그룹에 나타나도록 '다음' 그룹 이동을 반복."""
        hops = 0
        while hops < max_group_hops:
            scroll_to_pagination()
            link = find_page_link_in_container(target_page)
            if link:
                return True
            # 다음 그룹 이동 시도
            before_sig = get_first_list_href()
            print(f"타겟 {target_page}가 보이지 않아 '다음' 그룹 이동 시도")
            if not click_pagination_control("next"):
                break
            # 리스트 변화 대기 (첫 상품 href 변경 기준)
            for _ in range(10):
                time.sleep(0.5)
                after_sig = get_first_list_href()
                if before_sig and after_sig and before_sig != after_sig:
                    break
            hops += 1
        return False

    def get_current_page_number():
        # 0) 컨테이너의 pgn 값 파싱 시도
        pgn_val = get_pgn_from_container()
        if isinstance(pgn_val, int):
            return pgn_val
        selectors = [
            'a[aria-current="true"]',
            'button[aria-current="true"]',
            '[aria-current="page"]'
        ]
        # 컨테이너에서 role=menuitem + aria-current 우선 확인
        try:
            container = find_pagination_container()
            if container:
                node = container.query_selector("a[role='menuitem'][aria-current='true']")
                if node:
                    txt = (node.inner_text() or '').strip()
                    m = re.search(r'\d+', txt)
                    if m:
                        return int(m.group(0))
        except Exception:
            pass
        for selector in selectors:
            locator = page.locator(selector)
            try:
                locator.first.wait_for(state="attached", timeout=5000)
            except PlaywrightTimeoutError:
                continue

            try:
                text = locator.first.inner_text().strip()
            except PlaywrightTimeoutError:
                continue

            match = re.search(r'\d+', text)
            if match:
                return int(match.group())
        # 쿼리스트링에서 page 파라미터 추출 시도
        try:
            from urllib.parse import urlsplit, parse_qsl
            parts = urlsplit(page.url)
            query = dict(parse_qsl(parts.query, keep_blank_values=True))
            for key in ["page", "pageIndex", "pagingIndex", "pageNum", "p"]:
                if key in query:
                    try:
                        return int(re.findall(r"\d+", query[key])[0])
                    except Exception:
                        pass
        except Exception:
            pass
        print(f"현재 페이지 번호 탐색 실패 - URL: {page.url}")
        try:
            print("aria-current 후보:", page.locator('[aria-current]').all_inner_texts())
        except Exception:
            pass
        return None

    def find_page_link(target_page):
        # 1) 기존 네비게이션 링크(a[role=menuitem])에서 검색
        for link in page.query_selector_all('a[role="menuitem"]'):
            try:
                text = link.inner_text().strip()
            except PlaywrightTimeoutError:
                continue
            match = re.search(r'^\d+$', text)
            if match and int(text) == target_page:
                return link

        # 2) 접근성 역할 기반 탐색
        try:
            candidate = page.get_by_role("link", name=str(target_page))
            if candidate and candidate.count() > 0:
                return candidate.first
        except Exception:
            pass
        try:
            candidate = page.get_by_role("button", name=str(target_page))
            if candidate and candidate.count() > 0:
                return candidate.first
        except Exception:
            pass

        # 3) 포괄적 탐색(a, button)에서 텍스트가 숫자만이고 타겟과 일치하는 요소 선택
        for sel in ["a", "button"]:
            for link in page.query_selector_all(sel):
                try:
                    text = link.inner_text().strip()
                except Exception:
                    continue
                if not text:
                    continue
                if re.fullmatch(r"\d+", text) and int(text) == target_page:
                    return link

        return None

    def verify_first_product_on_page():
        try:
            # 첫 상품 링크 탐색
            page.wait_for_selector("a[href*='/products/']", timeout=10000)
            first_link = page.query_selector("a[href*='/products/']")
            if not first_link:
                print("검증 실패: 첫 상품 링크를 찾지 못했습니다.")
                return False
            href = first_link.get_attribute("href") or ""
            first_link.click()
            page.wait_for_load_state("load")
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError:
                pass
            final_url = page.url
            print(f"검증용 이동 URL: {final_url}")
            ok_url = True
            if verify_expected_url:
                ok_url = (verify_expected_url in final_url)
            ok_name = True
            if verify_expected_name:
                try:
                    content_text = page.inner_text("body")
                except Exception:
                    content_text = ""
                ok_name = (verify_expected_name in content_text)
            if ok_url and ok_name:
                print("검증 성공: 기대 URL/이름과 일치합니다.")
                return True
            else:
                print(f"검증 결과: URL일치={ok_url}, 이름일치={ok_name}")
                return False
        except Exception as exc:
            print(f"검증 중 예외: {exc}")
            return False

    def click_pagination_control(direction):
        labels = pagination_button_labels[direction]
        for label in labels:
            try:
                page.get_by_role("button", name=label).click(timeout=1500)
                page.wait_for_load_state("networkidle")
                time.sleep(1)
                return True
            except PlaywrightTimeoutError:
                continue
            except Exception:
                continue

        # 컨테이너 내부에서 '다음/이전' 우선 탐색
        container = find_pagination_container()
        if container:
            try:
                # role=button + aria-hidden=false 후보들 검사
                for node in container.query_selector_all("a[role='button'],button[role='button']"):
                    try:
                        hidden = node.get_attribute('aria-hidden')
                    except Exception:
                        hidden = None
                    if hidden == 'true':
                        continue
                    try:
                        t = (node.inner_text() or '').strip()
                    except Exception:
                        t = ''
                    if direction == 'next' and ("다음" in t or t in {"›", ">", "»"}):
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                    if direction == 'prev' and ("이전" in t or t in {"‹", "<", "«"}):
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                # 텍스트 기반 후보
                text_sel = (
                    "a:has-text('다음'),button:has-text('다음')" if direction == 'next' else "a:has-text('이전'),button:has-text('이전')"
                )
                cand = container.query_selector(text_sel)
                if cand:
                    try:
                        cand.scroll_into_view_if_needed()
                    except Exception:
                        pass
                    cand.click()
                    page.wait_for_load_state("networkidle")
                    time.sleep(1)
                    return True
                # 구조 기반 폴백: role=button 앵커 배열의 양 끝을 사용
                rb = container.query_selector_all("a[role='button'],button[role='button']")
                if rb:
                    try:
                        node = rb[-1] if direction == 'next' else rb[0]
                        try:
                            node.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                    except Exception:
                        pass
                # data-shp-contents-id 보유 요소 우선 클릭(네이버 특화)
                rb2 = container.query_selector_all("a[role='button'][data-shp-contents-id],button[role='button'][data-shp-contents-id]")
                if rb2:
                    try:
                        node = rb2[-1] if direction == 'next' else rb2[0]
                        try:
                            node.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                    except Exception:
                        pass
            except Exception:
                pass

        selector = (
            'a[role="button"][aria-hidden="false"]:last-child'
            if direction == "next"
            else 'a[role="button"][aria-hidden="false"]:first-child'
        )
        button = page.query_selector(selector)
        if button:
            button.click()
            page.wait_for_load_state("networkidle")
            time.sleep(1)
            return True

        return False

    def go_to_page_number(target_page):
        target_page = int(target_page)
        max_attempts = 40

        def list_signature():
            href = get_first_list_href()
            return href or ""

        attempt = 0
        while attempt < max_attempts:
            attempt += 1
            scroll_to_pagination()
            wait_pagination_ready(8000)

            cur = get_current_page_number()
            if cur == target_page:
                print(f"페이지 {target_page}에 이미 위치해 있습니다.")
                return True

            # ensure the current 10-page group exposes our target
            if not find_page_link_in_container(target_page):
                if target_page and cur and target_page > cur:
                    moved = click_pagination_control('next')
                else:
                    moved = click_pagination_control('prev')
                if not moved:
                    if not ensure_group_has_page(target_page):
                        print("타겟 페이지가 보이는 그룹으로 이동 실패")
                        continue
                before = list_signature()
                for _ in range(20):
                    time.sleep(0.5)
                    after = list_signature()
                    if before and after and before != after:
                        break

            link = find_page_link_in_container(target_page)
            if not link:
                continue

            try:
                link.scroll_into_view_if_needed()
            except Exception:
                pass

            before = list_signature()
            try:
                link.click()
            except Exception as e:
                print(f"숫자 링크 클릭 실패: {e}")
                continue

            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError:
                pass

            new_cur = get_current_page_number()
            if new_cur == target_page:
                print(f"페이지 {target_page}로 이동 완료, 현재 URL: {page.url}")
                return True

            for _ in range(20):
                time.sleep(0.5)
                after = list_signature()
                if before and after and before != after:
                    print(f"리스트 변경 감지로 페이지 {target_page} 이동 성공으로 간주")
                    return True

        print(f"페이지 {target_page} 이동 시도가 {max_attempts}회를 초과했습니다.")
        return False

    def url_prefix(u):
        filename = u.split("/")[-1]
        digits = "".join(ch for ch in filename if ch.isdigit())
        return digits[:3] if digits else filename[:3]

    prefixes = [url_prefix(url) for url in unique_urls]
    counts = Counter(prefixes)
    if not counts:
        return unique_urls, []

    most_common = counts.most_common(1)[0][0]
    common_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix == most_common]
    different_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix != most_common]

    return common_urls, different_urls


def content_crawl(page, product_code, element_selector):
    time.sleep(1)
    page.wait_for_load_state("load")

    if not element_selector:
        print("Invalid element selector. Moving to the next item.")
        return None

    element = page.query_selector(element_selector)
    if element is None:
        print("No valid element found for the selector. Moving to the next item.")
        return None

    content = element.inner_html()
    soup = BeautifulSoup(content, 'html.parser')

    if '계속됩니다' in soup.get_text():
        print("found 계속됩니다")
        return None

    css_link = soup.new_tag("link", rel="stylesheet",
                            href="https://static-resource-smartstore.pstatic.net/smartstore/p/static/20230630180923/common.css")
    if soup.head:
        soup.head.append(css_link)
    else:
        head_tag = soup.new_tag("head")
        head_tag.append(css_link)
        soup.insert(0, head_tag)

    for button in soup.find_all('button'):
        button.decompose()

    for img in soup.find_all('img', attrs={'data-src': True}):
        img['src'] = img['data-src']
        del img['data-src']

    target_url = "https://rapid-up.s3.ap-northeast-2.amazonaws.com/dev/gray-line.png"
    new_url = "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/gray-line.png"
    for img in soup.find_all('img', src=target_url):
        img['src'] = new_url

    text_to_remove = '* {text-align: center;}  #mycontents11 img{max-width: 100%;}'
    if text_to_remove in soup.get_text():
        soup = BeautifulSoup(str(soup).replace(text_to_remove, ''), 'html.parser')

    disallowed_attrs = ['area-hidden', 'data-linkdata', 'data-linktype', 'onclick', 'style', 'class']
    for attr in disallowed_attrs:
        for tag in soup.find_all(attrs={attr: True}):
            del tag[attr]

    for a in soup.find_all('a', attrs={'data-linkdata': True}):
        img = a.find('img')
        if img:
            src = img.get('src', '')
            data_src = img.get('data-src', '')
            if not src:
                src = data_src
                img['src'] = src
                if 'data-src' in img.attrs:
                    del img['data-src']

    for img in soup.find_all('img'):
        src = img.get('src', '')
        if src.startswith(('https://rapid-up.s3.ap-northeast-2.amazonaws.com', 'https://cdn.heyseller.kr', 'https://ai.esmplus.com/')):
            img.decompose()

    print(f"Number of images after all removals: {len(soup.find_all('img'))}")

    soup = insert_and_remove_images(soup)

    for img_tag in soup.find_all('img'):
        img_tag['style'] = 'display: block; margin-left: auto; margin-right: auto; margin-bottom: 10px;'

    for h1 in soup.find_all('h1'):
        h1['style'] = 'text-align: center; font-size: 30px; margin-bottom: 20px;'

    text_elements = ['p', 'div', 'span', 'li', 'a']
    for tag_name in text_elements:
        for element in soup.find_all(tag_name):
            existing_style = element.get('style', '')
            new_style = f"{existing_style}; text-align: center; font-size: 18px; margin-bottom: 30px;"
            element['style'] = new_style.strip()

    return pd.DataFrame({'Content': [str(soup)]})


def insert_and_remove_images(soup):
    img_srcs_to_insert = ['https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/top.png',
                          'https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/bottom.png',
                          'https://coudae.s3.ap-northeast-2.amazonaws.com/A00412936/cloud/7290.png']

    img_tag_top = soup.new_tag('img', src=img_srcs_to_insert[0],
                               style="display: block; margin-left: auto; margin-right: auto;")
    img_tag_middle = soup.new_tag('img', src=img_srcs_to_insert[2],
                                  style="display: block; margin-left: auto; margin-right: auto;")
    img_tag_bottom = soup.new_tag('img', src=img_srcs_to_insert[1],
                                  style="display: block; margin-left: auto; margin-right: auto;")

    first_tag = next(soup.children)
    last_tag = next(reversed(soup.contents))

    first_tag.insert_before(img_tag_top)
    last_tag.insert_after(img_tag_bottom)

    img_srcs_to_remove = ['', '']

    for img_src in img_srcs_to_remove:
        img_to_remove = soup.find_all('img', attrs={'src': img_src})
        for img in img_to_remove:
            img.decompose()

    return soup


def return_shipping_fee(total_price):
    fee = total_price * 0.25
    if fee > 200000:
        fee = 200000
    return fee


def title_edit(title):
    title_split = title.split(' ')
    title_split = list(dict.fromkeys(title_split))
    if len(title_split) >= 2:
        title_split[-1], title_split[-2] = title_split[-2], title_split[-1]
    title = ' '.join(title_split)
    return title


def get_product_data(page, product, i, num_products):
    title, price, product_url, product_code = extract_product_details(product)

    title = title.replace('\xa0', ' ')

    print(f"Product {i + 1}/{num_products}: {title}, {price} won, {product_url}")

    product_page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(product_page)
    product_page.goto(product_url)
    product_page.wait_for_load_state("load")
    try:
        product_page.wait_for_load_state("networkidle", timeout=10000)
    except PlaywrightTimeoutError:
        pass

    scripts = product_page.query_selector_all('script')

    if not NAVER_CATEGORY_PATH.exists():
        raise FileNotFoundError(f"카테고리 파일을 찾을 수 없습니다: {NAVER_CATEGORY_PATH}")
    category_df = pd.read_excel(NAVER_CATEGORY_PATH, header=None)
    small_category_dict = pd.Series(category_df[0].values, index=category_df[3]).to_dict()
    tiny_category_dict = pd.Series(category_df[0].values, index=category_df[4]).to_dict()

    category = None
    for script in scripts:
        script_content = script.inner_text()
        if "category" in script_content:
            try:
                json_data = json.loads(script_content)
            except Exception:
                continue
            if 'category' in json_data:
                category = json_data['category']
                break

    if category is not None:
        print(f"Category: {category}")
        category_list = category.split(">")
        large_category = category_list[0].strip()
        medium_category = category_list[1].strip()
        small_category = category_list[2].strip() if len(category_list) > 2 else None
        tiny_category = category_list[3].strip() if len(category_list) > 3 else None
    else:
        category_list = []
        large_category = None
        medium_category = None
        small_category = None
        tiny_category = None

    if tiny_category is not None:
        smallest_category = tiny_category
        smallest_category_type = 'tiny'
        naver_category_number = tiny_category_dict.get(tiny_category)
    elif small_category is not None:
        smallest_category = small_category
        smallest_category_type = 'small'
        naver_category_number = small_category_dict.get(small_category)
    else:
        smallest_category = None
        smallest_category_type = None
        naver_category_number = None

    print(
        f"Smallest Category('{smallest_category_type}') : {smallest_category}, Naver category number: {naver_category_number}")

    options = option_crawl(product_page)
    print("Options:", options)

    common_urls, different_urls = image_crawl(product_page)
    print(f"common_urls: {common_urls}")

    main_image = None
    other_images = []

    if common_urls:
        main_image = common_urls[0].replace('?type=m510', '')
        other_images = [url.replace('?type=m510', '') for url in common_urls[1:]]
        print(f"other_images: {other_images}")
    else:
        print("No common images found")
        try:
            image_element = page.wait_for_selector('xpath=//*[@id="content"]/div/div[2]/div[1]/div[1]/div[1]/img', timeout=2000)
            image_url = image_element.get_attribute("src")
            main_image = image_url.replace('?type=m510', '')
        except Exception:
            print("No main image found")

    print("Main image:", main_image)
    print("Other images:", other_images)

    print("URLs not starting with most common three digits:")
    for url in different_urls:
        print(url)

    element_selector = find_content_element(product_page, product_code)
    content = content_crawl(product_page, product_code, element_selector)
    shipping_fee = original_shipping_fee(product_page)

    def to_int(value):
        if value in (None, "N/A"):
            return 0
        digits = re.sub(r"[^\d]", "", str(value))
        return int(digits) if digits else 0

    shipping_fee_int = to_int(shipping_fee)
    price_int = to_int(price)
    total_price = price_int + shipping_fee_int

    if isinstance(content, str):
        content_df = pd.DataFrame({'Content': [content]})
    else:
        content_df = content

    product_df = pd.DataFrame({
        'Product': [title],
        'Price': [price],
        'Shipping_Fee': [shipping_fee_int],
        'Total_Price': [total_price],
        'Main_Image': [main_image],
        'Other_Images': [other_images],
        'Options': [options],
        'Product_URL': [product_url],
        'Naver_Category_Number': [naver_category_number]
    })

    product_df = pd.concat([product_df, content_df], axis=1)
    product_page.close()

    return product_df


def write_to_excel(df, excel_path, seen_urls):
    book = load_workbook(excel_path)
    sheet = book['일괄등록']

    b_start_row = c_start_row = e_start_row = h_start_row = ad_start_row = r_start_row = s_start_row = t_start_row = i_start_row = 3
    ap_start_row = aq_start_row = 3

    # 데이터가 없으면 템플릿만 저장하고 조용히 반환
    if df is None or len(df) == 0:
        print("DataFrame이 비어 있어 엑셀 기록을 생략합니다.")
        book.save(excel_path)
        if os.name != "nt":
            print(f"Excel file saved to {excel_path}. (empty dataset)")
        return

    for i, item in enumerate(df['Naver_Category_Number'], start=b_start_row):
        sheet['B' + str(i)] = item
    for j, item in enumerate(df['Product'], start=c_start_row):
        sheet['C' + str(j)] = item
    for k, (product_price, shipping_fee) in enumerate(zip(df['Price'], df['Shipping_Fee']), start=e_start_row):
        total_price = float(product_price.replace(',', '')) + shipping_fee
        selling_price = total_price - 0.01 * total_price
        selling_price_rounded = round(selling_price / 100.0) * 100.0
        sheet['E' + str(k)] = selling_price_rounded

    for l, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['H' + str(l)] = "조합형"
    for m in range(ad_start_row, ad_start_row + len(df)):
        selling_price_rounded = sheet['E' + str(m)].value

        # 바젤마켓 분기
        if selling_price_rounded <= 20000:
            ad_value = 2903608
        elif 20001 <= selling_price_rounded <= 30000:
            ad_value = 2904260
        elif 30001 <= selling_price_rounded <= 40000:
            ad_value = 2904261
        elif 40001 <= selling_price_rounded <= 60000:
            ad_value = 2904262
        elif 60001 <= selling_price_rounded <= 80000:
            ad_value = 2904268
        elif 80001 <= selling_price_rounded <= 100000:
            ad_value = 2904272
        elif 100001 <= selling_price_rounded <= 150000:
            ad_value = 2904276
        elif 150001 <= selling_price_rounded <= 400000:
            ad_value = 2904278
        elif 400001 <= selling_price_rounded <= 600000:
            ad_value = 2904279
        elif 600001 <= selling_price_rounded <= 1000000:
            ad_value = 2904281
        elif 1000001 <= selling_price_rounded <= 9999999:
            ad_value = 2904284

        sheet['AD' + str(m)] = ad_value

    for row, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['U' + str(row)] = "상세페이지 참조"
        sheet['V' + str(row)] = "상세페이지 참조"
        sheet['Y' + str(row)] = "0200037"
        sheet['Z' + str(row)] = "구매대행"
        sheet['AZ' + str(row)] = "010-3973-3119"
        sheet['BA' + str(row)] = "본문 안내문 참조"

    for l, item in enumerate(df['Options'], start=r_start_row):
        option_titles = []
        option_prices = []
        option_categories = []
        for key in item:
            option_categories.append(key)
            if '하위옵션제목' in item[key]:
                option_titles.append(', '.join(item[key]['하위옵션제목']))
            if '하위옵션가격' in item[key]:
                option_prices.append(', '.join(map(str, item[key]['하위옵션가격'])))
        sheet['I' + str(l)] = '\n'.join(option_categories)
        sheet['J' + str(l)] = '\n'.join(option_titles)
        sheet['K' + str(l)] = '\n'.join(option_prices)

    if 'Options' in df.columns:
        for row in range(h_start_row, h_start_row + len(df)):
            item_options = df.at[row - h_start_row, 'Options']
            option_prices = []
            for option in item_options.values():
                if '하위옵션가격' in option:
                    option_prices.extend(option['하위옵션가격'])

            if option_prices:
                num_prices = len(option_prices)
                l_values = ', '.join(['99'] * num_prices)
                sheet['L' + str(row)] = l_values
            else:
                sheet['L' + str(row)] = "99"
    else:
        for row in range(h_start_row, h_start_row + len(df)):
            sheet['L' + str(row)] = "99"

    for n, item in enumerate(df['Main_Image'], start=r_start_row):
        sheet['R' + str(n)] = item
    for o, item in enumerate(df['Other_Images'], start=s_start_row):
        if item is not None:
            sheet['S' + str(o)] = "\n".join(str(img) for img in item if img is not None)
        else:
            sheet['S' + str(o)] = ""

    if 'Content' in df.columns:
        for p, item in enumerate(df['Content'], start=t_start_row):
            if isinstance(item, float):
                item = str(item)
            print(f"Row {p}, Content: {item[:100]}")
            sheet['T' + str(p)] = item
    else:
        print("No 'Content' column found in DataFrame")

    for q, url in enumerate(df['Product_URL'], start=i_start_row):
        product_code = url.split('/')[-1]
        try:
            selling_code = str(int(product_code) * 2)
        except ValueError:
            selling_code = str(random.randint(10000000, 99999999)) + 'R'
        sheet['A' + str(q)] = selling_code
    for r, total_price in enumerate(df['Total_Price'], start=ap_start_row):
        return_fee = return_shipping_fee(total_price)
        return_fee_rounded = round(return_fee / 100.0) * 100
        sheet['AP' + str(r)] = return_fee_rounded
        sheet['AQ' + str(r)] = return_fee_rounded * 2

    book.save(excel_path)

    # 뒤쪽 빈 행 제거
    book = load_workbook(excel_path)
    sheet = book['일괄등록']
    for i in range(3, sheet.max_row + 1):
        if not sheet['A' + str(i)].value:
            sheet.delete_rows(i, sheet.max_row - i + 1)
            break
    book.save(excel_path)

    if os.name == "nt":
        os.system(f'start "" "excel.exe" "{excel_path}"')
    else:
        print(f"Excel file saved to {excel_path}. Automatic Excel launch is skipped on non-Windows platforms.")


def write_to_excel2(df, excel_path2):
    df2 = pd.DataFrame({
        'Product_URL': df['Product_URL'],
        'Numbering': range(1, len(df) + 1),
        'Product_Title': df['Product'],
        'Product_Price': df['Price'],
        'Shipping_Fee': df['Shipping_Fee']
    })
    with pd.ExcelWriter(excel_path2) as writer:
        df2.to_excel(writer, index=False)


# 실행부: 항상 로컬 브라우저를 실행 (Windows 우선)
if CRAWLER_DRY_RUN:
    print("CRAWLER_DRY_RUN=1 플래그로 인해 Playwright 크롤링 본동작을 생략합니다.")
    sys.stdout = original
    f.close()
    sys.exit(0)

with sync_playwright() as p:
    browser_name = os.getenv("PLAYWRIGHT_BROWSER", "chromium").lower()
    if browser_name not in {"chromium", "firefox", "webkit"}:
        browser_name = "chromium"

    headless_mode = os.getenv("PLAYWRIGHT_HEADLESS", "0").lower() in {"1", "true", "yes"}

    if browser_name == "chromium":
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-features=NetworkService",
            "--disable-web-security",
            "--disable-dev-shm-usage",
            "--disable-accelerated-2d-canvas",
            "--disable-gpu",
        ]
        browser = p.chromium.launch(headless=headless_mode, args=launch_args)
        context = browser.new_context()
    elif browser_name == "firefox":
        browser = p.firefox.launch(headless=headless_mode)
        context = browser.new_context()
    else:
        browser = p.webkit.launch(headless=headless_mode)
        context = browser.new_context()

    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(context)

    df = pd.DataFrame(columns=['Product', 'Price', 'Product_URL'])
    output_folder = SCRIPT_DIR / 'output'
    read_excel_path = output_folder / 'ExcelSaveTemplate_230109.xlsx'
    seen_urls = set()

    product_list_crawl(context, df, read_excel_path, seen_urls)
    try:
        context.close()
    finally:
        browser.close()

sys.stdout = original
f.close()