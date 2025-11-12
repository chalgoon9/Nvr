from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
try:
    from playwright_stealth import Stealth
except ImportError:
    Stealth = None
from collections import Counter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd
import random
import time
import shutil
import re
import os
import json
import requests
import sys
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
# ==== DEBUG / LOGGING UTILITIES ====

# ==== PLAYWRIGHT EVENT WIRING ====
def wire_browser_events(browser):
    try:
        browser.on("disconnected", lambda _: log("BROWSER EVENT: disconnected"))
    except Exception:
        pass

def wire_context_events(context):
    try:
        context.on("close", lambda _: log("CONTEXT EVENT: close"))
        context.on("page", lambda p: log(f"CONTEXT EVENT: new page -> {getattr(p, 'url', None)}"))
    except Exception:
        pass

def wire_page_events(page):
    try:
        page.on("close", lambda: log("PAGE EVENT: close"))
        page.on("crash", lambda: log("PAGE EVENT: crash"))
        page.on("domcontentloaded", lambda: log(f"PAGE EVENT: domcontentloaded url={page.url}"))
        page.on("load", lambda: log(f"PAGE EVENT: load url={page.url}"))
        page.on("console", lambda msg: log(f"PAGE CONSOLE: {msg.type} {msg.text}"))
        page.on("pageerror", lambda err: log(f"PAGE ERROR: {err}"))
        page.on("request", lambda req: log(f"REQ  {req.method} {req.url}"))
        page.on("requestfailed", lambda req: log(f"REQ FAIL {req.method} {req.url} -> {req.failure}"))
        page.on("response", lambda res: log(f"RES  {res.status} {res.url}"))
        page.on("dialog", lambda dlg: (log(f"DIALOG: {dlg.type} {dlg.message}"), dlg.dismiss()))
    except Exception:
        pass
# ==================================
DEBUG = bool(int(os.getenv("DEBUG", "1") or "1"))
DEBUG_HOLD = int(os.getenv("DEBUG_HOLD_SEC", "0") or "0")
DEBUG_DIR = os.getenv("DEBUG_DIR", "debug")
os.makedirs(DEBUG_DIR, exist_ok=True)

def log(msg: str):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        with open(os.path.join(DEBUG_DIR, "run.log"), "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass
# ===================================




# 페이지당 최대 크롤링 상품 수 (0이면 제한 없음)
MAX_PRODUCTS_PER_PAGE = int(os.getenv("MAX_PRODUCTS_PER_PAGE", "0") or 0)


# .env 로더 (python-dotenv 미설치 시 최소 파서)
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None


def fallback_load_dotenv(dotenv_path):
    path = Path(dotenv_path)
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        return False
    except OSError as exc:
        print(f".env 파일을 읽는 중 오류가 발생했습니다: {exc}")
        return False

    loaded = False
    saw_assignments = False
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        saw_assignments = True
        key = key.strip()
        value = value.strip().strip("\"'")
        if not key:
            continue
        if key not in os.environ:
            os.environ[key] = value
            loaded = True
    return loaded or saw_assignments


SCRIPT_DIR = Path(__file__).resolve().parent
LOG_FILE = SCRIPT_DIR / "log.txt"


def resolve_category_path() -> Path:
    # 우선 로컬 경로, 없으면 상위 디렉터리(AGENTS.md 가이드에 맞춤)
    local = SCRIPT_DIR / "naver_category.xlsx"
    parent = SCRIPT_DIR.parent / "naver_category.xlsx"
    if local.exists():
        return local
    if parent.exists():
        return parent
    # 마지막으로 로컬 경로를 반환(실패 시 런타임에서 에러 메시지 제공)
    return local


NAVER_CATEGORY_PATH = resolve_category_path()
CRAWLER_DRY_RUN = os.getenv("CRAWLER_DRY_RUN", "0").lower() in {"1", "true", "yes"}

# 선택 페이지만 크롤링하는 디버그용 옵션(예: CRAWL_ONLY_PAGES="51,59").
# 지정되지 않으면 기존 범위(global_start_page~global_last_page) 전체를 처리합니다.
_only_pages_raw = os.getenv("CRAWL_ONLY_PAGES", "").strip()
CRAWL_ONLY_PAGES = None
if _only_pages_raw:
    try:
        parts = re.split(r"[\s,;]+", _only_pages_raw)
        CRAWL_ONLY_PAGES = [int(p) for p in parts if p]
    except Exception as exc:
        print(f"CRAWL_ONLY_PAGES 파싱 실패({_only_pages_raw}): {exc}")
        CRAWL_ONLY_PAGES = None

# 페이지 번호 이동 시 URL 쿼리 파라미터로 강제 점프 시도 여부
# 기본값은 비활성화(네이버는 URL 파라미터만으로 DOM이 바뀌지 않는 경우가 많음)
PAGE_JUMP_BY_QUERY = os.getenv("PAGE_JUMP_BY_QUERY", "0").lower() in {"1", "true", "yes"}

# 페이지 이동 과정 스크린샷 저장(디버깅 용도)
# 기본값 비활성화: 요청에 따라 캡처 중단
PAGINATION_DEBUG_SHOTS = os.getenv("PAGINATION_DEBUG_SHOTS", "0").lower() in {"1", "true", "yes"}

# 페이지네이션 전략: auto(기본) | next_only('다음'만 반복)
PAGINATION_STRATEGY = os.getenv("PAGINATION_STRATEGY", "auto").lower().strip()

def debug_shot(page, label):
    if not PAGINATION_DEBUG_SHOTS:
        return
    try:
        ts = time.strftime("%Y%m%d_%H%M%S")
        dbg = (SCRIPT_DIR / "debug")
        dbg.mkdir(exist_ok=True)
        path = dbg / f"pagination_{ts}_{label}.png"
        page.screenshot(path=str(path), full_page=True)
        print(f"Saved screenshot: {path}")
    except Exception as exc:
        print(f"Failed to take screenshot({label}): {exc}")

STEALTH_HELPER = Stealth() if Stealth is not None else None
if STEALTH_HELPER is None:
    print(
        "playwright_stealth 모듈에서 Stealth 클래스를 불러오지 못했습니다. "
        "탐지 회피 스크립트가 적용되지 않으니 chromium 환경에서는 추가 점검이 필요합니다."
    )


if load_dotenv:
    load_dotenv()
else:
    _ = fallback_load_dotenv(SCRIPT_DIR / ".env")


def first_available(node, selectors):
    for selector in selectors:
        try:
            element = node.query_selector(selector)
        except PlaywrightTimeoutError:
            continue
        if element:
            return element
    return None


def find_elements(page, selectors):
    for selector in selectors:
        try:
            elements = page.query_selector_all(selector)
        except PlaywrightTimeoutError:
            continue
        if elements:
            print(f"Selector '{selector}' matched {len(elements)} elements.")
            return elements
    return []


def save_debug_snapshot(page, prefix):
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    path = (SCRIPT_DIR / "debug")
    path.mkdir(exist_ok=True)
    out = path / f"{prefix}_{timestamp}.html"
    try:
        out.write_text(page.content(), encoding="utf-8")
        print(f"Saved debug snapshot: {out}")
    except Exception as exc:
        print(f"Failed to save debug snapshot: {exc}")


def extract_price_from_text(raw_text):
    match = re.search(r"([\d,]+)\s*원", raw_text)
    if match:
        return match.group(1).replace("\u200b", "").strip()
    digits = re.sub(r"[^\d]", "", raw_text)
    if not digits:
        return "N/A"
    try:
        return "{:,}".format(int(digits))
    except ValueError:
        return "N/A"


def update_query_params(url, **params):
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({key: value for key, value in params.items() if value is not None})
    new_query = urlencode(query, doseq=True)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))


class Tee(object):
    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()

    def flush(self):
        for f in self.files:
            f.flush()


f = LOG_FILE.open('w', encoding='utf-8')
original = sys.stdout
sys.stdout = Tee(sys.stdout, f)


base_url = "https://smartstore.naver.com"


def product_list_crawl(context, df, read_excel_path, seen_urls):
    page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(page)

    raw_url = 'https://smartstore.naver.com/joypapa_/category/ALL?st=RECENT&dt=BIG_IMAGE&size=80'
    original_url = update_query_params(raw_url, page=None)
    try:

        log(f"GOTO: page.goto(original_url)")

        page.goto(original_url)

        log(f"GOTO DONE url={page.url}")

    except Exception as e:

        log(f"GOTO FAILED: {e}")

        try:

            page.screenshot(path=os.path.join(DEBUG_DIR, "goto_failure.png"))

            log("Saved goto_failure.png")

        except Exception:

            pass

        if DEBUG_HOLD:

            log(f"Holding for {DEBUG_HOLD} seconds because of failure.")

            page.wait_for_timeout(DEBUG_HOLD * 1000)

    page.wait_for_load_state("load")
    page.wait_for_load_state("networkidle")

    global_start_page = 51
    global_last_page = 59
    # 디버그: 특정 페이지만 요청된 경우 범위를 해당 값으로 축소
    if CRAWL_ONLY_PAGES:
        try:
            global_start_page = min(CRAWL_ONLY_PAGES)
            global_last_page = max(CRAWL_ONLY_PAGES)
        except Exception:
            pass
    shopname = raw_url.split('/')[3]
    shopnumber = raw_url.split('/')[5].split('?')[0]

    home_dir = Path.home()
    output_folder = home_dir / 'Desktop' / 'excel_output'
    output_folder.mkdir(parents=True, exist_ok=True)

    pagination_button_labels = {
        "next": ["다음", "다음 페이지", "다음페이지", ">"],
        "prev": ["이전", "이전 페이지", "이전페이지", "<"]
    }

    # 검증 모드: 특정 페이지의 첫 상품을 열어 기대 URL/이름 확인
    verify_target_page = None
    verify_expected_url = (os.getenv("VERIFY_FIRST_PRODUCT_URL") or "").strip() or None
    verify_expected_name = (os.getenv("VERIFY_FIRST_PRODUCT_NAME") or "").strip() or None
    try:
        verify_target_page = int(os.getenv("VERIFY_TARGET_PAGE", "") or 0) or None
    except Exception:
        verify_target_page = None

    def scroll_to_pagination():
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        except PlaywrightTimeoutError:
            pass
        time.sleep(0.8)
        # 하단 스크롤 후 간단 스크린샷
        debug_shot(page, "scrolled_bottom")


def wait_pagination_ready(timeout_ms=10000):
    """Wait until a pagination container with clickable items appears."""
    end_ts = time.time() + timeout_ms/1000.0
    while time.time() < end_ts:
        container = find_pagination_container()
        if container:
            try:
                # numbers or buttons present?
                has_items = container.locator("a,button,[role='link'],[role='button']").count()
                if has_items and has_items > 0:
                    return True
            except Exception:
                pass
        time.sleep(0.2)
    return False

    def get_first_list_href():
        try:
            el = page.query_selector("a[href*='/products/']")
            if el:
                href = el.get_attribute("href")
                return href
        except Exception:
            pass
        return None


def find_pagination_container():
    """Return a *Locator* to the pagination container, or None."""
    candidates = [
        "div[data-shp-area='list.pgn'][role='menubar']",
        "div[data-shp-contents-type='pgn'][role='menubar']",
        "div[data-shp-area-id='pgn'][role='menubar']",
        "nav[aria-label*='페이지']",
        "nav[aria-label*='pagination']",
        "nav[role='navigation']",
        "div[class*='Pagination']",
        "div[class*='paginate']",
        "div[class*='paging']",
    ]
    for sel in candidates:
        try:
            loc = page.locator(sel)
            if loc.count() > 0:
                return loc.first
        except Exception:
            continue
    return None


def get_visible_page_numbers():
    """Return sorted list of page numbers currently rendered inside the pagination component."""
    container = find_pagination_container()
    if not container:
        return []
    numbers = []
    try:
        items = container.locator("a,button,span,li")
        cnt = items.count()
        for i in range(cnt):
            try:
                t = (items.nth(i).inner_text() or "").strip()
            except Exception:
                continue
            if t.isdigit():
                n = int(t)
                if n not in numbers:
                    numbers.append(n)
    except Exception:
        pass
    numbers.sort()
    return numbers


def get_pgn_from_container():
    """Parse pgn/page value from container's data attribute if present."""
    try:
        container = find_pagination_container()
        if not container:
            return None
        # We need the attribute; use first element handle for attribute read
        handle = container.element_handle()
        if not handle:
            return None
        raw = handle.get_attribute("data-shp-filter_con")
        if not raw:
            return None
        raw = raw.replace('&quot;', '"')
        data = json.loads(raw)
        if isinstance(data, list):
            for item in data:
                if isinstance(item, dict) and item.get('key') in {'pgn', 'page', 'pageNum'}:
                    val = item.get('value')
                    try:
                        return int(re.findall(r"\d+", str(val))[0])
                    except Exception:
                        pass
    except Exception:
        return None
    return None


def find_page_link_in_container(target_page):
    container = find_pagination_container()
    if container is None:
        return None
    # Prefer ARIA role lookup
    target_name = str(target_page)
    for role in ("link", "button"):
        try:
            loc = container.get_by_role(role, name=target_name)
            if loc and loc.count() > 0:
                return loc.first
        except Exception:
            pass
    # Fallback: any anchor/button with exact numeric text
    try:
        nodes = container.locator("a,button")
        cnt = nodes.count()
        for i in range(cnt):
            try:
                t = (nodes.nth(i).inner_text() or "").strip()
            except Exception:
                continue
            if t.isdigit() and int(t) == target_page:
                return nodes.nth(i)
    except Exception:
        pass
    return None


def ensure_group_has_page(target_page, max_group_hops=30):
    """Navigate page-number groups (10-by-10) until the group containing target_page is visible."""
    hops = 0
    while hops < max_group_hops:
        scroll_to_pagination()
        wait_pagination_ready(6000)

        # Already visible?
        link = find_page_link_in_container(target_page)
        if link:
            return True

        visible_numbers = get_visible_page_numbers()
        if visible_numbers:
            min_v, max_v = min(visible_numbers), max(visible_numbers)
            if min_v <= target_page <= max_v:
                # sometimes the DOM needs a tick to attach handlers
                time.sleep(0.2)
                hops += 1
                continue
            direction = 'next' if target_page > max_v else 'prev'
        else:
            # cannot read; try advancing forward
            direction = 'next'

        print(f"[ensure_group_has_page] visible={visible_numbers or 'none'}, click {direction}")
        if not click_pagination_control(direction):
            print("[ensure_group_has_page] could not click pagination control; giving up.")
            return False

        hops += 1

    print(f"[ensure_group_has_page] exceeded max hops while seeking page {target_page}.")
    return False


def get_current_page_number():
    # 0) Parse from container data if present
    pgn_val = get_pgn_from_container()
    if isinstance(pgn_val, int):
        return pgn_val

    # 1) aria-current within container
    try:
        container = find_pagination_container()
        if container:
            cur = container.locator("a[aria-current='true'],button[aria-current='true'],[aria-current='page']")
            if cur.count() > 0:
                txt = (cur.first.inner_text() or '').strip()
                m = re.search(r"\d+", txt)
                if m:
                    return int(m.group(0))
    except Exception:
        pass

    # 2) Global aria-current (fallback)
    try:
        cur = page.locator("[aria-current]")
        if cur.count() > 0:
            txt = (cur.first.inner_text() or '').strip()
            m = re.search(r"\d+", txt)
            if m:
                return int(m.group(0))
    except Exception:
        pass

    # 3) URL params (very weak on Naver, but keep as last resort)
    try:
        parts = urlsplit(page.url)
        query = dict(parse_qsl(parts.query, keep_blank_values=True))
        for key in ["page", "pageIndex", "pagingIndex", "pageNum", "p"]:
            if key in query:
                try:
                    return int(re.findall(r"\d+", query[key])[0])
                except Exception:
                    pass
    except Exception:
        pass

    return None

    def find_page_link(target_page):
        # 1) 기존 네비게이션 링크(a[role=menuitem])에서 검색
        for link in page.query_selector_all('a[role="menuitem"]'):
            try:
                text = link.inner_text().strip()
            except PlaywrightTimeoutError:
                continue
            match = re.search(r'^\d+$', text)
            if match and int(text) == target_page:
                return link

        # 2) 접근성 역할 기반 탐색
        try:
            candidate = page.get_by_role("link", name=str(target_page))
            if candidate and candidate.count() > 0:
                return candidate.first
        except Exception:
            pass
        try:
            candidate = page.get_by_role("button", name=str(target_page))
            if candidate and candidate.count() > 0:
                return candidate.first
        except Exception:
            pass

        # 3) 포괄적 탐색(a, button)에서 텍스트가 숫자만이고 타겟과 일치하는 요소 선택
        for sel in ["a", "button"]:
            for link in page.query_selector_all(sel):
                try:
                    text = link.inner_text().strip()
                except Exception:
                    continue
                if not text:
                    continue
                if re.fullmatch(r"\d+", text) and int(text) == target_page:
                    return link

        return None

    def verify_first_product_on_page():
        try:
            # 첫 상품 링크 탐색
            page.wait_for_selector("a[href*='/products/']", timeout=10000)
            first_link = page.query_selector("a[href*='/products/']")
            if not first_link:
                print("검증 실패: 첫 상품 링크를 찾지 못했습니다.")
                return False
            href = first_link.get_attribute("href") or ""
            first_link.click()
            page.wait_for_load_state("load")
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError:
                pass
            final_url = page.url
            print(f"검증용 이동 URL: {final_url}")
            ok_url = True
            if verify_expected_url:
                ok_url = (verify_expected_url in final_url)
            ok_name = True
            if verify_expected_name:
                try:
                    content_text = page.inner_text("body")
                except Exception:
                    content_text = ""
                ok_name = (verify_expected_name in content_text)
            if ok_url and ok_name:
                print("검증 성공: 기대 URL/이름과 일치합니다.")
                return True
            else:
                print(f"검증 결과: URL일치={ok_url}, 이름일치={ok_name}")
                return False
        except Exception as exc:
            print(f"검증 중 예외: {exc}")
            return False


def click_pagination_control(direction):
    """Click 'next' or 'prev' control inside pagination.
    Returns True if we think the page list changed.
    """
    before_sig = get_first_list_href()
    labels_next = ["다음", "다음 페이지", "다음페이지", ">", "›", "»"]
    labels_prev = ["이전", "이전 페이지", "이전페이지", "<", "‹", "«"]
    label_list = labels_next if direction == 'next' else labels_prev

    def try_click(loc):
        try:
            # bring into view and click
            loc.scroll_into_view_if_needed(timeout=2000)
        except Exception:
            pass
        try:
            loc.click(timeout=3000, force=False)
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            # wait for product list signature to change
            for _ in range(20):
                time.sleep(0.3)
                after = get_first_list_href()
                if before_sig and after and before_sig != after:
                    return True
            return True  # be optimistic if we at least clicked something
        except Exception:
            return False

    # 1) Try within pagination container
    container = find_pagination_container()
    if container:
        # Try role-based buttons/links by name
        for label in label_list:
            for role in ("button", "link"):
                try:
                    loc = container.get_by_role(role, name=label)
                    if loc and loc.count() > 0:
                        if try_click(loc.first):
                            return True
                except Exception:
                    continue
        # 2) Heuristic: pick last/first visible role=button if names failed
        try:
            buttons = container.locator("a[role='button'],button[role='button'],a,button")
            count = buttons.count()
            if count > 0:
                loc = buttons.nth(count - 1) if direction == 'next' else buttons.nth(0)
                if try_click(loc):
                    return True
        except Exception:
            pass

    # 3) Fallback: search whole page
    for label in label_list:
        for role in ("button", "link"):
            try:
                loc = page.get_by_role(role, name=label)
                if loc and loc.count() > 0:
                    if try_click(loc.first):
                        return True
            except Exception:
                continue

    return False


def go_to_page_number(target_page):
    """Robustly navigate to the given page using only UI pagination (no URL params)."""
    target_page = int(target_page)
    print(f"[go_to_page_number] → {target_page}")
    # Try to ensure the page group is visible
    if not ensure_group_has_page(target_page):
        print(f"[go_to_page_number] group with {target_page} not visible; attempt direct advance.")
    # Click the number if visible
    link = find_page_link_in_container(target_page)
    if link:
        before = get_first_list_href()
        try:
            link.scroll_into_view_if_needed(timeout=2000)
        except Exception:
            pass
        try:
            debug_shot(page, f"goto_{target_page}_before")
            link.click(timeout=4000)
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            debug_shot(page, f"goto_{target_page}_after")
        except Exception as exc:
            print(f"[go_to_page_number] click number failed: {exc}")
        # verify
        num = get_current_page_number()
        if num == target_page:
            print(f"[go_to_page_number] now at page {target_page} (aria/attr). URL={page.url}")
            return True
        # verify by list signature change
        after = get_first_list_href()
        if before and after and before != after:
            print(f"[go_to_page_number] list changed; assume reached page {target_page}")
            return True

    # As a fallback, step with next/prev until we reach target
    cur = get_current_page_number()
    if cur is None:
        # assume starting at 1 if unknown
        cur = 1
    direction = 'next' if target_page > cur else 'prev'
    steps = abs(target_page - cur) + 15  # generous headroom
    for _ in range(min(300, steps)):
        if cur == target_page:
            print(f"[go_to_page_number] arrived at {target_page} by stepping.")
            return True
        if not click_pagination_control(direction):
            print(f"[go_to_page_number] failed stepping {direction}.")
            return False
        new_cur = get_current_page_number()
        cur = new_cur if new_cur is not None else (cur + (1 if direction=='next' else -1))
    return get_current_page_number() == target_page

    def url_prefix(u):
        filename = u.split("/")[-1]
        digits = "".join(ch for ch in filename if ch.isdigit())
        return digits[:3] if digits else filename[:3]

    prefixes = [url_prefix(url) for url in unique_urls]
    counts = Counter(prefixes)
    if not counts:
        return unique_urls, []

    most_common = counts.most_common(1)[0][0]
    common_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix == most_common]
    different_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix != most_common]

    return common_urls, different_urls


def content_crawl(page, product_code, element_selector):
    time.sleep(1)
    page.wait_for_load_state("load")

    if not element_selector:
        print("Invalid element selector. Moving to the next item.")
        return None

    element = page.query_selector(element_selector)
    if element is None:
        print("No valid element found for the selector. Moving to the next item.")
        return None

    content = element.inner_html()
    soup = BeautifulSoup(content, 'html.parser')

    if '계속됩니다' in soup.get_text():
        print("found 계속됩니다")
        return None

    css_link = soup.new_tag("link", rel="stylesheet",
                            href="https://static-resource-smartstore.pstatic.net/smartstore/p/static/20230630180923/common.css")
    if soup.head:
        soup.head.append(css_link)
    else:
        head_tag = soup.new_tag("head")
        head_tag.append(css_link)
        soup.insert(0, head_tag)

    for button in soup.find_all('button'):
        button.decompose()

    for img in soup.find_all('img', attrs={'data-src': True}):
        img['src'] = img['data-src']
        del img['data-src']

    target_url = "https://rapid-up.s3.ap-northeast-2.amazonaws.com/dev/gray-line.png"
    new_url = "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/gray-line.png"
    for img in soup.find_all('img', src=target_url):
        img['src'] = new_url

    text_to_remove = '* {text-align: center;}  #mycontents11 img{max-width: 100%;}'
    if text_to_remove in soup.get_text():
        soup = BeautifulSoup(str(soup).replace(text_to_remove, ''), 'html.parser')

    disallowed_attrs = ['area-hidden', 'data-linkdata', 'data-linktype', 'onclick', 'style', 'class']
    for attr in disallowed_attrs:
        for tag in soup.find_all(attrs={attr: True}):
            del tag[attr]

    for a in soup.find_all('a', attrs={'data-linkdata': True}):
        img = a.find('img')
        if img:
            src = img.get('src', '')
            data_src = img.get('data-src', '')
            if not src:
                src = data_src
                img['src'] = src
                if 'data-src' in img.attrs:
                    del img['data-src']

    for img in soup.find_all('img'):
        src = img.get('src', '')
        if src.startswith(('https://rapid-up.s3.ap-northeast-2.amazonaws.com', 'https://cdn.heyseller.kr', 'https://ai.esmplus.com/')):
            img.decompose()

    print(f"Number of images after all removals: {len(soup.find_all('img'))}")

    soup = insert_and_remove_images(soup)

    for img_tag in soup.find_all('img'):
        img_tag['style'] = 'display: block; margin-left: auto; margin-right: auto; margin-bottom: 10px;'

    for h1 in soup.find_all('h1'):
        h1['style'] = 'text-align: center; font-size: 30px; margin-bottom: 20px;'

    text_elements = ['p', 'div', 'span', 'li', 'a']
    for tag_name in text_elements:
        for element in soup.find_all(tag_name):
            existing_style = element.get('style', '')
            new_style = f"{existing_style}; text-align: center; font-size: 18px; margin-bottom: 30px;"
            element['style'] = new_style.strip()

    return pd.DataFrame({'Content': [str(soup)]})


def insert_and_remove_images(soup):
    img_srcs_to_insert = ['https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/top.png',
                          'https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/bottom.png',
                          'https://coudae.s3.ap-northeast-2.amazonaws.com/A00412936/cloud/7290.png']

    img_tag_top = soup.new_tag('img', src=img_srcs_to_insert[0],
                               style="display: block; margin-left: auto; margin-right: auto;")
    img_tag_middle = soup.new_tag('img', src=img_srcs_to_insert[2],
                                  style="display: block; margin-left: auto; margin-right: auto;")
    img_tag_bottom = soup.new_tag('img', src=img_srcs_to_insert[1],
                                  style="display: block; margin-left: auto; margin-right: auto;")

    first_tag = next(soup.children)
    last_tag = next(reversed(soup.contents))

    first_tag.insert_before(img_tag_top)
    last_tag.insert_after(img_tag_bottom)

    img_srcs_to_remove = ['', '']

    for img_src in img_srcs_to_remove:
        img_to_remove = soup.find_all('img', attrs={'src': img_src})
        for img in img_to_remove:
            img.decompose()

    return soup


def return_shipping_fee(total_price):
    fee = total_price * 0.25
    if fee > 200000:
        fee = 200000
    return fee


def title_edit(title):
    title_split = title.split(' ')
    title_split = list(dict.fromkeys(title_split))
    if len(title_split) >= 2:
        title_split[-1], title_split[-2] = title_split[-2], title_split[-1]
    title = ' '.join(title_split)
    return title


def get_product_data(page, product, i, num_products):
    title, price, product_url, product_code = extract_product_details(product)

    title = title.replace('\xa0', ' ')

    print(f"Product {i + 1}/{num_products}: {title}, {price} won, {product_url}")

    product_page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(product_page)
    product_page.goto(product_url)
    product_page.wait_for_load_state("load")
    try:
        product_page.wait_for_load_state("networkidle", timeout=10000)
    except PlaywrightTimeoutError:
        pass

    scripts = product_page.query_selector_all('script')

    if not NAVER_CATEGORY_PATH.exists():
        raise FileNotFoundError(f"카테고리 파일을 찾을 수 없습니다: {NAVER_CATEGORY_PATH}")
    category_df = pd.read_excel(NAVER_CATEGORY_PATH, header=None)
    small_category_dict = pd.Series(category_df[0].values, index=category_df[3]).to_dict()
    tiny_category_dict = pd.Series(category_df[0].values, index=category_df[4]).to_dict()

    category = None
    for script in scripts:
        script_content = script.inner_text()
        if "category" in script_content:
            try:
                json_data = json.loads(script_content)
            except Exception:
                continue
            if 'category' in json_data:
                category = json_data['category']
                break

    if category is not None:
        print(f"Category: {category}")
        category_list = category.split(">")
        large_category = category_list[0].strip()
        medium_category = category_list[1].strip()
        small_category = category_list[2].strip() if len(category_list) > 2 else None
        tiny_category = category_list[3].strip() if len(category_list) > 3 else None
    else:
        category_list = []
        large_category = None
        medium_category = None
        small_category = None
        tiny_category = None

    if tiny_category is not None:
        smallest_category = tiny_category
        smallest_category_type = 'tiny'
        naver_category_number = tiny_category_dict.get(tiny_category)
    elif small_category is not None:
        smallest_category = small_category
        smallest_category_type = 'small'
        naver_category_number = small_category_dict.get(small_category)
    else:
        smallest_category = None
        smallest_category_type = None
        naver_category_number = None

    print(
        f"Smallest Category('{smallest_category_type}') : {smallest_category}, Naver category number: {naver_category_number}")

    options = option_crawl(product_page)
    print("Options:", options)

    common_urls, different_urls = image_crawl(product_page)
    print(f"common_urls: {common_urls}")

    main_image = None
    other_images = []

    if common_urls:
        main_image = common_urls[0].replace('?type=m510', '')
        other_images = [url.replace('?type=m510', '') for url in common_urls[1:]]
        print(f"other_images: {other_images}")
    else:
        print("No common images found")
        try:
            image_element = page.wait_for_selector('xpath=//*[@id="content"]/div/div[2]/div[1]/div[1]/div[1]/img', timeout=2000)
            image_url = image_element.get_attribute("src")
            main_image = image_url.replace('?type=m510', '')
        except Exception:
            print("No main image found")

    print("Main image:", main_image)
    print("Other images:", other_images)

    print("URLs not starting with most common three digits:")
    for url in different_urls:
        print(url)

    element_selector = find_content_element(product_page, product_code)
    content = content_crawl(product_page, product_code, element_selector)
    shipping_fee = original_shipping_fee(product_page)

    def to_int(value):
        if value in (None, "N/A"):
            return 0
        digits = re.sub(r"[^\d]", "", str(value))
        return int(digits) if digits else 0

    shipping_fee_int = to_int(shipping_fee)
    price_int = to_int(price)
    total_price = price_int + shipping_fee_int

    if isinstance(content, str):
        content_df = pd.DataFrame({'Content': [content]})
    else:
        content_df = content

    product_df = pd.DataFrame({
        'Product': [title],
        'Price': [price],
        'Shipping_Fee': [shipping_fee_int],
        'Total_Price': [total_price],
        'Main_Image': [main_image],
        'Other_Images': [other_images],
        'Options': [options],
        'Product_URL': [product_url],
        'Naver_Category_Number': [naver_category_number]
    })

    product_df = pd.concat([product_df, content_df], axis=1)
    product_page.close()

    return product_df


def write_to_excel(df, excel_path, seen_urls):
    book = load_workbook(excel_path)
    sheet = book['일괄등록']

    b_start_row = c_start_row = e_start_row = h_start_row = ad_start_row = r_start_row = s_start_row = t_start_row = i_start_row = 3
    ap_start_row = aq_start_row = 3

    # 데이터가 없으면 템플릿만 저장하고 조용히 반환
    if df is None or len(df) == 0:
        print("DataFrame이 비어 있어 엑셀 기록을 생략합니다.")
        book.save(excel_path)
        if os.name != "nt":
            print(f"Excel file saved to {excel_path}. (empty dataset)")
        return

    for i, item in enumerate(df['Naver_Category_Number'], start=b_start_row):
        sheet['B' + str(i)] = item
    for j, item in enumerate(df['Product'], start=c_start_row):
        sheet['C' + str(j)] = item
    for k, (product_price, shipping_fee) in enumerate(zip(df['Price'], df['Shipping_Fee']), start=e_start_row):
        total_price = float(product_price.replace(',', '')) + shipping_fee
        selling_price = total_price - 0.01 * total_price
        selling_price_rounded = round(selling_price / 100.0) * 100.0
        sheet['E' + str(k)] = selling_price_rounded

    for l, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['H' + str(l)] = "조합형"
    for m in range(ad_start_row, ad_start_row + len(df)):
        selling_price_rounded = sheet['E' + str(m)].value

        # 바젤마켓 분기
        if selling_price_rounded <= 20000:
            ad_value = 2903608
        elif 20001 <= selling_price_rounded <= 30000:
            ad_value = 2904260
        elif 30001 <= selling_price_rounded <= 40000:
            ad_value = 2904261
        elif 40001 <= selling_price_rounded <= 60000:
            ad_value = 2904262
        elif 60001 <= selling_price_rounded <= 80000:
            ad_value = 2904268
        elif 80001 <= selling_price_rounded <= 100000:
            ad_value = 2904272
        elif 100001 <= selling_price_rounded <= 150000:
            ad_value = 2904276
        elif 150001 <= selling_price_rounded <= 400000:
            ad_value = 2904278
        elif 400001 <= selling_price_rounded <= 600000:
            ad_value = 2904279
        elif 600001 <= selling_price_rounded <= 1000000:
            ad_value = 2904281
        elif 1000001 <= selling_price_rounded <= 9999999:
            ad_value = 2904284

        sheet['AD' + str(m)] = ad_value

    for row, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['U' + str(row)] = "상세페이지 참조"
        sheet['V' + str(row)] = "상세페이지 참조"
        sheet['Y' + str(row)] = "0200037"
        sheet['Z' + str(row)] = "구매대행"
        sheet['AZ' + str(row)] = "010-3973-3119"
        sheet['BA' + str(row)] = "본문 안내문 참조"

    for l, item in enumerate(df['Options'], start=r_start_row):
        option_titles = []
        option_prices = []
        option_categories = []
        for key in item:
            option_categories.append(key)
            if '하위옵션제목' in item[key]:
                option_titles.append(', '.join(item[key]['하위옵션제목']))
            if '하위옵션가격' in item[key]:
                option_prices.append(', '.join(map(str, item[key]['하위옵션가격'])))
        sheet['I' + str(l)] = '\n'.join(option_categories)
        sheet['J' + str(l)] = '\n'.join(option_titles)
        sheet['K' + str(l)] = '\n'.join(option_prices)

    if 'Options' in df.columns:
        for row in range(h_start_row, h_start_row + len(df)):
            item_options = df.at[row - h_start_row, 'Options']
            option_prices = []
            for option in item_options.values():
                if '하위옵션가격' in option:
                    option_prices.extend(option['하위옵션가격'])

            if option_prices:
                num_prices = len(option_prices)
                l_values = ', '.join(['99'] * num_prices)
                sheet['L' + str(row)] = l_values
            else:
                sheet['L' + str(row)] = "99"
    else:
        for row in range(h_start_row, h_start_row + len(df)):
            sheet['L' + str(row)] = "99"

    for n, item in enumerate(df['Main_Image'], start=r_start_row):
        sheet['R' + str(n)] = item
    for o, item in enumerate(df['Other_Images'], start=s_start_row):
        if item is not None:
            sheet['S' + str(o)] = "\n".join(str(img) for img in item if img is not None)
        else:
            sheet['S' + str(o)] = ""

    if 'Content' in df.columns:
        for p, item in enumerate(df['Content'], start=t_start_row):
            if isinstance(item, float):
                item = str(item)
            print(f"Row {p}, Content: {item[:100]}")
            sheet['T' + str(p)] = item
    else:
        print("No 'Content' column found in DataFrame")

    for q, url in enumerate(df['Product_URL'], start=i_start_row):
        product_code = url.split('/')[-1]
        try:
            selling_code = str(int(product_code) * 2)
        except ValueError:
            selling_code = str(random.randint(10000000, 99999999)) + 'R'
        sheet['A' + str(q)] = selling_code
    for r, total_price in enumerate(df['Total_Price'], start=ap_start_row):
        return_fee = return_shipping_fee(total_price)
        return_fee_rounded = round(return_fee / 100.0) * 100
        sheet['AP' + str(r)] = return_fee_rounded
        sheet['AQ' + str(r)] = return_fee_rounded * 2

    book.save(excel_path)

    # 뒤쪽 빈 행 제거
    book = load_workbook(excel_path)
    sheet = book['일괄등록']
    for i in range(3, sheet.max_row + 1):
        if not sheet['A' + str(i)].value:
            sheet.delete_rows(i, sheet.max_row - i + 1)
            break
    book.save(excel_path)

    if os.name == "nt":
        os.system(f'start "" "excel.exe" "{excel_path}"')
    else:
        print(f"Excel file saved to {excel_path}. Automatic Excel launch is skipped on non-Windows platforms.")


def write_to_excel2(df, excel_path2):
    df2 = pd.DataFrame({
        'Product_URL': df['Product_URL'],
        'Numbering': range(1, len(df) + 1),
        'Product_Title': df['Product'],
        'Product_Price': df['Price'],
        'Shipping_Fee': df['Shipping_Fee']
    })
    with pd.ExcelWriter(excel_path2) as writer:
        df2.to_excel(writer, index=False)


# 실행부: 항상 로컬 브라우저를 실행 (Windows 우선)
if CRAWLER_DRY_RUN:
    print("CRAWLER_DRY_RUN=1 플래그로 인해 Playwright 크롤링 본동작을 생략합니다.")
    sys.stdout = original
    f.close()
    sys.exit(0)

with sync_playwright() as p:

log("PLAYWRIGHT START")

# Chromium selection & headless from env
headless_env = os.getenv("HEADLESS", "0")
headless = False if headless_env in ("0","false","False","no","NO") else True
slowmo_ms = int(os.getenv("SLOW_MO_MS", "0") or 0)

try:
    log(f"Launching chromium headless={headless} slow_mo={slowmo_ms}")
    browser = p.chromium.launch(headless=headless, slow_mo=slowmo_ms)
    wire_browser_events(browser)
except Exception as e:
    log(f"Browser launch failed: {e}")
    raise

try:
    ctx_opts = {}
    # Optional: record video for the first page
    record_video = os.getenv("PW_RECORD_VIDEO", "1")
    if record_video not in ("0","false","False","no","NO"):
        ctx_opts['record_video_dir'] = DEBUG_DIR
        ctx_opts['record_video_size'] = {'width': 1280, 'height': 720}
    context = browser.new_context(**ctx_opts)
    wire_context_events(context)
    start_tracing(context)
    log("Context created.")
except Exception as e:
    log(f"Context creation failed: {e}")
    raise

try:
    page = context.new_page()
    wire_page_events(page)
    log("New page created and wired.")
except Exception as e:
    log(f"Page creation failed: {e}")
    raise

    browser_name = os.getenv("PLAYWRIGHT_BROWSER", "chromium").lower()
    if browser_name not in {"chromium", "firefox", "webkit"}:
        browser_name = "chromium"

    headless_mode = os.getenv("PLAYWRIGHT_HEADLESS", "0").lower() in {"1", "true", "yes"}

    if browser_name == "chromium":
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-features=NetworkService",
            "--disable-web-security",
            "--disable-dev-shm-usage",
            "--disable-accelerated-2d-canvas",
            "--disable-gpu",
        ]
        browser = p.chromium.launch(headless=headless_mode, args=launch_args)
        context = browser.new_context()
    elif browser_name == "firefox":
        browser = p.firefox.launch(headless=headless_mode)
        context = browser.new_context()
    else:
        browser = p.webkit.launch(headless=headless_mode)
        context = browser.new_context()

    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(context)

    df = pd.DataFrame(columns=['Product', 'Price', 'Product_URL'])
    output_folder = SCRIPT_DIR / 'output'
    read_excel_path = output_folder / 'ExcelSaveTemplate_230109.xlsx'
    seen_urls = set()

    product_list_crawl(context, df, read_excel_path, seen_urls)
    try:
        _debug_footer(context, page)

        context.close()
    finally:
        browser.close()

sys.stdout = original
f.close()



def start_tracing(context):
    try:
        context.tracing.start(screenshots=True, snapshots=True, sources=True)
        log("Tracing started.")
    except Exception as e:
        log(f"Tracing start failed: {e}")

def stop_tracing(context, path="debug/trace.zip"):
    try:
        context.tracing.stop(path=path)
        log(f"Tracing saved to {path}")
    except Exception as e:
        log(f"Tracing stop failed: {e}")



def _debug_footer(context, page):
    try:
        if page and hasattr(page, "video") and page.video:
            try:
                vp = page.video.path()
                log(f"Video path: {vp}")
            except Exception as e:
                log(f"Video path retrieval failed: {e}")
    except Exception:
        pass
    try:
        stop_tracing(context, path=os.path.join(DEBUG_DIR, "trace.zip"))
    except Exception:
        pass
    if DEBUG_HOLD:
        log(f"DEBUG_HOLD active. Keeping page open for {DEBUG_HOLD} seconds.")
        try:
            page.wait_for_timeout(DEBUG_HOLD * 1000)
        except Exception:
            time.sleep(DEBUG_HOLD)

