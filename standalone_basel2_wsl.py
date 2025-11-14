from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
try:
    from playwright_stealth import Stealth
except ImportError:
    Stealth = None
from collections import Counter
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook
import pandas as pd
import random
import time
import shutil
import re
#from oracle_cloud import upload_to_oracle_cloud
import os
import json
import requests
import base64
import socket
import sys
import platform
import subprocess
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode
import struct

def running_on_wsl():
    if "WSL_DISTRO_NAME" in os.environ:
        return True
    try:
        return "microsoft" in platform.uname().release.lower()
    except Exception:
        return False


def _private_ipv4(ip):
    try:
        parts = [int(p) for p in ip.split(".")]
        if len(parts) != 4:
            return False
        a, b, *_ = parts
        return (
            a == 10
            or (a == 172 and 16 <= b <= 31)
            or (a == 192 and b == 168)
        )
    except Exception:
        return False


def _wsl_default_gateway():
    """Detect default gateway IP inside WSL by parsing /proc/net/route."""
    try:
        with open("/proc/net/route", "r", encoding="utf-8") as fp:
            lines = fp.read().splitlines()
        for line in lines[1:]:
            cols = line.split()  # Iface  Destination  Gateway ...
            if len(cols) >= 3 and cols[1] == "00000000":
                gw_hex = cols[2]
                try:
                    gw_int = int(gw_hex, 16)
                    gw_bytes = struct.pack("<L", gw_int)
                    gw_ip = socket.inet_ntoa(gw_bytes)
                    return gw_ip
                except Exception:
                    continue
    except OSError:
        pass
    return None


def detect_windows_host_from_wsl():
    """Return Windows host IP from WSL with multiple strategies.

    Priority order:
    1) env WSL_CDP_HOST/PLAYWRIGHT_CDP_HOST
    2) /proc/net/route default gateway (if private range)
    3) resolv.conf nameserver (if private range)
    """
    if not running_on_wsl():
        return None

    override = (os.getenv("WSL_CDP_HOST") or os.getenv("PLAYWRIGHT_CDP_HOST") or "").strip()
    if override:
        return override

    gw = _wsl_default_gateway()
    if gw and _private_ipv4(gw):
        return gw

    resolv_path = Path("/etc/resolv.conf")
    try:
        for line in resolv_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("nameserver"):
                parts = line.split()
                if len(parts) >= 2 and parts[1] not in {"127.0.0.1", "::1"}:
                    if _private_ipv4(parts[1]):
                        return parts[1]
    except OSError:
        pass
    return None


def determine_default_cdp_origin():
    wsl_host = detect_windows_host_from_wsl()
    if wsl_host:
        return f"http://{wsl_host}:9222"
    return "http://127.0.0.1:9222"


DEFAULT_CONNECT_URL = determine_default_cdp_origin()
_DEFAULT_PARTS = urlsplit(DEFAULT_CONNECT_URL)
DEFAULT_CDP_HOST = _DEFAULT_PARTS.hostname or "127.0.0.1"
DEFAULT_CDP_PORT = _DEFAULT_PARTS.port or 9222


def locate_cmd_invocation():
    if os.name == "nt":
        return ["cmd", "/c"]
    if running_on_wsl():
        candidates = [
            "/mnt/c/Windows/System32/cmd.exe",
            "/mnt/c/Windows/system32/cmd.exe",
            "/mnt/c/windows/System32/cmd.exe",
            "/mnt/c/windows/system32/cmd.exe",
        ]
        for candidate in candidates:
            if Path(candidate).exists():
                return [candidate, "/c"]
    return None


def to_windows_path(path):
    path = Path(path).resolve()
    as_str = str(path)
    if os.name == "nt":
        return as_str
    if running_on_wsl():
        if as_str.startswith("/mnt/") and len(as_str) > 6:
            drive_letter = as_str[5].upper()
            remainder = as_str[7:].replace("/", "\\")
            return f"{drive_letter}:\\" + remainder
    return as_str


def normalize_cdp_url(raw_url):
    if not raw_url:
        return None
    candidate = raw_url.strip()
    if not candidate:
        return None
    if "://" not in candidate:
        candidate = f"http://{candidate}"
    parts = urlsplit(candidate)
    scheme = parts.scheme.lower()
    hostname = parts.hostname or DEFAULT_CDP_HOST
    if hostname in {"0.0.0.0", "*"}:
        hostname = DEFAULT_CDP_HOST
    port = parts.port or DEFAULT_CDP_PORT
    netloc = f"{hostname}:{port}" if port else hostname
    normalized_scheme = "https" if scheme in {"https", "wss"} else "http"
    return urlunsplit((normalized_scheme, netloc, "", "", ""))

# ?섏씠吏??理쒕? ?щ·留??곹뭹 ??(0?대㈃ ?쒗븳 ?놁쓬)
MAX_PRODUCTS_PER_PAGE = int(os.getenv("MAX_PRODUCTS_PER_PAGE", "0") or 0)

## ?쒖옉

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None


def fallback_load_dotenv(dotenv_path):
    """Minimal .env parser used when python-dotenv is unavailable."""
    path = Path(dotenv_path)
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        return False
    except OSError as exc:
        print(f".env ?뚯씪???쎈뒗 以??ㅻ쪟媛 諛쒖깮?덉뒿?덈떎: {exc}")
        return False

    loaded = False
    saw_assignments = False
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        saw_assignments = True
        key = key.strip()
        value = value.strip().strip("\"'")  # drop simple quoting
        if not key:
            continue
        # ?대? ?섍꼍蹂?섍? 吏?뺣릺???덈떎硫???뼱?곗? ?딅뒗??
        if key not in os.environ:
            os.environ[key] = value
            loaded = True
    return loaded or saw_assignments

SCRIPT_DIR = Path(__file__).resolve().parent
LOG_FILE = SCRIPT_DIR / "log.txt"
NAVER_CATEGORY_PATH = SCRIPT_DIR / "naver_category.xlsx"
AUTO_LAUNCH_CHROME_DEVTOOLS = os.getenv("AUTO_LAUNCH_CHROME_DEVTOOLS", "1").lower() not in {"0", "false", "no"}
REQUIRE_CDP_CONNECTION = os.getenv("REQUIRE_CDP_CONNECTION", "0").lower() in {"1", "true", "yes"}
CRAWLER_DRY_RUN = os.getenv("CRAWLER_DRY_RUN", "0").lower() in {"1", "true", "yes"}
# Windows-only 媛꾩냼 紐⑤뱶: CDP ?곌껐 諛?諛곗튂 ?ㅽ뻾 紐⑤몢 ?앸왂?섍퀬 濡쒖뺄 Chromium??吏곸젒 ?ㅽ뻾
FORCE_LOCAL_PLAYWRIGHT = os.getenv("FORCE_LOCAL_PLAYWRIGHT", "0").lower() in {"1", "true", "yes"}

STEALTH_HELPER = Stealth() if Stealth is not None else None
if STEALTH_HELPER is None:
    print(
        "playwright_stealth 紐⑤뱢?먯꽌 Stealth ?대옒?ㅻ? 遺덈윭?ㅼ? 紐삵뻽?듬땲?? "
        "?먯? ?뚰뵾 ?ㅽ겕由쏀듃媛 ?곸슜?섏? ?딆쑝??chromium ?섍꼍?먯꽌??異붽? ?먭????꾩슂?⑸땲??"
    )


if load_dotenv:
    load_dotenv()
else:
    dotenv_loaded = fallback_load_dotenv(Path(__file__).with_name(".env"))
    if not dotenv_loaded:
        print(
            "python-dotenv 紐⑤뱢??李얠쓣 ???놁뼱 .env ?뚯씪???먮룞?쇰줈 濡쒕뱶?섏? 紐삵뻽?듬땲?? "
            f"'pip install python-dotenv'濡?紐⑤뱢???ㅼ튂?섍굅??PLAYWRIGHT_CONNECT_URL ?섍꼍蹂???? {DEFAULT_CONNECT_URL})瑜?吏곸젒 吏?뺥븯?몄슂."
        )


def verify_cdp_endpoint(url, timeout=5):
    """?ъ쟾 ?묒냽 ?щ? ?뺤씤 (/json/version)."""
    probe_url = url.rstrip("/") + "/json/version"
    try:
        # ?쇰? ?섍꼍?먯꽌 Host ?ㅻ뜑媛 localhost???뚮쭔 ?묐떟?섎뒗 ?щ?媛 ?덉뼱 ?ㅼ젙
        response = requests.get(probe_url, timeout=timeout, headers={"Host": "localhost"})
        response.raise_for_status()
        return True
    except requests.RequestException as exc:
        print(
            "Chrome ?먭꺽 ?붾쾭源??몄뀡???곌껐?섏? 紐삵뻽?듬땲??\n"
            f"?뺤씤 URL: {probe_url}\n"
            "Chrome??--remote-debugging-port=9222 ?듭뀡?쇰줈 ?ㅽ뻾 以묒씤吏, "
            "WSL?먯꽌 ?대떦 ?ы듃濡??묎렐 媛?ν븳吏 ?ㅼ떆 ?뺤씤??二쇱꽭??\n"
            f"?곸꽭 ?ㅻ쪟: {exc}"
        )
        return False


def _extract_first_json_object(text):
    """Return first JSON object substring found in text, or None.

    Useful to parse Windows cmd.exe output that may prepend UNC warnings to JSON.
    """
    if not text:
        return None
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start : end + 1]
    return None

base_url = "https://smartstore.naver.com"
DEBUG_DIR = SCRIPT_DIR / "debug"
DEBUG_DIR.mkdir(exist_ok=True)


def maybe_launch_chrome_devtools():
    if FORCE_LOCAL_PLAYWRIGHT:
        print("FORCE_LOCAL_PLAYWRIGHT=1 ?대?濡?Chrome DevTools ?먮룞 ?ㅽ뻾??嫄대꼫?곷땲??")
        return
    if CRAWLER_DRY_RUN:
        print("CRAWLER_DRY_RUN=1 ?대?濡?Chrome DevTools ?먮룞 ?ㅽ뻾??嫄대꼫?곷땲??")
        return

    if not AUTO_LAUNCH_CHROME_DEVTOOLS:
        print("AUTO_LAUNCH_CHROME_DEVTOOLS=0 ?대?濡?Chrome DevTools ?먮룞 ?ㅽ뻾??嫄대꼫?곷땲??")
        return

    batch_path = SCRIPT_DIR / "start_chrome_dev.bat"
    if not batch_path.exists():
        print(f"{batch_path} ?뚯씪??李얠쓣 ???놁뼱 Chrome DevTools ?먮룞 ?ㅽ뻾???앸왂?⑸땲??")
        return

    cmd_parts = locate_cmd_invocation()
    if cmd_parts is None:
        print("cmd.exe 瑜?李얠? 紐삵빐 Chrome DevTools ?먮룞 ?ㅽ뻾???앸왂?⑸땲?? Windows ?섍꼍?먯꽌 吏곸젒 諛곗튂?뚯씪???ㅽ뻾??二쇱꽭??")
        return

    if os.name == "nt":
        batch_argument = str(batch_path)
    elif running_on_wsl():
        batch_argument = to_windows_path(batch_path)
    else:
        batch_argument = str(batch_path)

    try:
        subprocess.run(
            [*cmd_parts, batch_argument],
            check=True,
            capture_output=True,
            text=True,
            timeout=30,
        )
    except FileNotFoundError as exc:
        print(f"Chrome DevTools 諛곗튂 ?ㅽ뻾 ?ㅽ뙣 (cmd.exe 誘몃컻寃?: {exc}")
    except PermissionError as exc:
        print(
            "Chrome DevTools 諛곗튂 ?ㅽ뻾 沅뚰븳 ?ㅻ쪟媛 諛쒖깮?덉뒿?덈떎. "
            "WSL interop ?ㅼ젙 ?먮뒗 /mnt/c ?쒕씪?대툕 ?ㅽ뻾 沅뚰븳???뺤씤?섍퀬 吏곸젒 諛곗튂?뚯씪???ㅽ뻾??二쇱꽭??\n"
            f"?곸꽭: {exc}"
        )
    except subprocess.CalledProcessError as exc:
        print(
            "Chrome DevTools 諛곗튂 ?ㅽ뻾???ㅽ뙣?덉뒿?덈떎.\n"
            f"returncode: {exc.returncode}\n"
            f"stdout: {exc.stdout}\n"
            f"stderr: {exc.stderr}"
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout or ""
        stderr = exc.stderr or ""
        print(
            "Chrome DevTools 諛곗튂 ?ㅽ뻾??30珥??덉뿉 醫낅즺?섏? ?딆븘 ?щ·?ш? ?湲곕? 以묐떒?⑸땲?? "
            "Chrome 李쎌씠 ?대? ???덈떎硫?洹몃?濡?吏꾪뻾?⑸땲??\n"
            f"stdout: {stdout}\n"
            f"stderr: {stderr}"
        )
    except OSError as exc:
        print(f"Chrome DevTools 諛곗튂 ?ㅽ뻾 以??????녿뒗 OS ?ㅻ쪟媛 諛쒖깮?덉뒿?덈떎: {exc}")


def first_available(node, selectors):
    """Return the first element that matches one of the selectors."""
    for selector in selectors:
        try:
            element = node.query_selector(selector)
        except PlaywrightTimeoutError:
            continue
        if element:
            return element
    return None


def find_elements(page, selectors):
    """Return the first non-empty element list found via the selector list."""
    for selector in selectors:
        try:
            elements = page.query_selector_all(selector)
        except PlaywrightTimeoutError:
            continue
        if elements:
            print(f"Selector '{selector}' matched {len(elements)} elements.")
            return elements
    return []


def save_debug_snapshot(page, prefix):
    """Persist current HTML to debug/ for selector troubleshooting."""
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    path = DEBUG_DIR / f"{prefix}_{timestamp}.html"
    try:
        path.write_text(page.content(), encoding="utf-8")
        print(f"Saved debug snapshot: {path}")
    except Exception as exc:
        print(f"Failed to save debug snapshot: {exc}")


def extract_price_from_text(raw_text):
    match = re.search(r"([\d,]+)\s*??, raw_text)
    if match:
        return match.group(1).replace("\u200b", "").strip()
    digits = re.sub(r"[^\d]", "", raw_text)
    if not digits:
        return "N/A"
    try:
        return "{:,}".format(int(digits))
    except ValueError:
        return "N/A"


def update_query_params(url, **params):
    """Add or replace query parameters in a URL."""
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({key: value for key, value in params.items() if value is not None})
    new_query = urlencode(query, doseq=True)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))

class Tee(object):
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()
    def flush(self):
        for f in self.files:
            f.flush()

f = LOG_FILE.open('w', encoding='utf-8')
original = sys.stdout
sys.stdout = Tee(sys.stdout, f)

def product_list_crawl(context, df, read_excel_path, seen_urls):
    page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(page)

    raw_url = 'https://smartstore.naver.com/joypapa_/category/ALL?st=RECENT&dt=BIG_IMAGE&size=80'
    original_url = update_query_params(raw_url, page=None)
    page.goto(original_url)
    page.wait_for_load_state("load")
    page.wait_for_load_state("networkidle")

    global_start_page = 51
    global_last_page = 59
    shopname = raw_url.split('/')[3]
    shopnumber = raw_url.split('/')[5].split('?')[0]
    home_dir = Path.home()
    output_folder = home_dir / 'Desktop' / 'excel_output'
    output_folder.mkdir(parents=True, exist_ok=True)

    pagination_button_labels = {
        "next": ["?ㅼ쓬", "?ㅼ쓬 ?섏씠吏", "?ㅼ쓬?섏씠吏", ">"],
        "prev": ["?댁쟾", "?댁쟾 ?섏씠吏", "?댁쟾?섏씠吏", "<"]
    }

    def scroll_to_pagination():
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        except PlaywrightTimeoutError:
            pass
        time.sleep(0.8)

    def get_current_page_number():
        selectors = [
            'a[aria-current="true"]',
            'button[aria-current="true"]',
            '[aria-current="page"]'
        ]
        for selector in selectors:
            locator = page.locator(selector)
            try:
                locator.first.wait_for(state="attached", timeout=5000)
            except PlaywrightTimeoutError:
                continue

            try:
                text = locator.first.inner_text().strip()
            except PlaywrightTimeoutError:
                continue

            match = re.search(r'\d+', text)
            if match:
                return int(match.group())
        print(f"?꾩옱 ?섏씠吏 踰덊샇 ?먯깋 ?ㅽ뙣 - URL: {page.url}")
        try:
            print("aria-current ?꾨낫:", page.locator('[aria-current]').all_inner_texts())
        except Exception:
            pass
        return None

    def find_page_link(target_page):
        for link in page.query_selector_all('a[role="menuitem"]'):
            try:
                text = link.inner_text().strip()
            except PlaywrightTimeoutError:
                continue
            match = re.search(r'^\d+$', text)
            if match and int(text) == target_page:
                return link
        return None

    def click_pagination_control(direction):
        labels = pagination_button_labels[direction]
        for label in labels:
            try:
                page.get_by_role("button", name=label).click(timeout=1500)
                page.wait_for_load_state("networkidle")
                time.sleep(1)
                return True
            except PlaywrightTimeoutError:
                continue
            except Exception:
                continue

        selector = (
            'a[role="button"][aria-hidden="false"]:last-child'
            if direction == "next"
            else 'a[role="button"][aria-hidden="false"]:first-child'
        )
        button = page.query_selector(selector)
        if button:
            button.click()
            page.wait_for_load_state("networkidle")
            time.sleep(1)
            return True

        return False

    def go_to_page_number(target_page):
        attempt = 0
        max_attempts = 30
        while attempt < max_attempts:
            attempt += 1
            scroll_to_pagination()
            current_page_num = get_current_page_number()

            if current_page_num == target_page:
                print(f"?섏씠吏 {target_page}???대? ?꾩튂???덉뒿?덈떎.")
                return True

            page_link = find_page_link(target_page)
            if page_link:
                try:
                    page_link.scroll_into_view_if_needed()
                except PlaywrightTimeoutError:
                    pass
                time.sleep(0.3)
                page_link.click()
                page.wait_for_load_state("networkidle")
                time.sleep(1)
                if get_current_page_number() == target_page:
                    print(f"?섏씠吏 {target_page}濡??대룞 ?꾨즺, ?꾩옱 URL: {page.url}")
                    return True
                continue

            if current_page_num is None:
                print("?꾩옱 ?섏씠吏 踰덊샇瑜??뺤씤?????놁뼱 ?ㅼ떆 ?쒕룄?⑸땲??")
                page.wait_for_timeout(1000)
                continue

            direction = "next" if target_page > current_page_num else "prev"
            print(f"?섏씠吏 {target_page} ?대룞???꾪빐 {direction} 踰꾪듉 ?대┃ ?쒕룄 (?꾩옱 {current_page_num}).")
            if not click_pagination_control(direction):
                print(f"{direction} 踰꾪듉??李얠쓣 ???놁뒿?덈떎.")
                return False

        print(f"?섏씠吏 {target_page} ?대룞 ?쒕룄媛 {max_attempts}??珥덇낵濡??ㅽ뙣?덉뒿?덈떎.")
        return False

    for start_page in range(global_start_page, global_last_page + 1, 10):
        last_page = min(start_page + 9, global_last_page)
        write_excel_path = output_folder / f'dolce_{shopname}_{shopnumber}_{start_page}_{last_page}.xlsx'
        shutil.copy(read_excel_path, write_excel_path)

        if not go_to_page_number(start_page):
            print(f"?섏씠吏 {start_page} ?대룞???ㅽ뙣?덉뒿?덈떎. ?ㅼ쓬 洹몃９?쇰줈 ?섏뼱媛묐땲??")
            continue

        for page_number in range(start_page, last_page + 1):
            if not go_to_page_number(page_number):
                print(f"?섏씠吏 {page_number} ?대룞???ㅽ뙣?섏뿬 嫄대꼫?곷땲??")
                continue
            df, _ = crawl_page(page, df, seen_urls)
            print(f"Completed page {page_number}")

        write_to_excel(df, write_excel_path, seen_urls)
        write_to_excel2(
            df,
            output_folder / f'dolce_{shopname}_{shopnumber}_{start_page}_{last_page}_second.xlsx'
        )
        print(f"Processed pages {start_page} to {last_page}")

    page.close()


def find_content_element(page, product_code):
    time.sleep(1)

    content_selectors = [
        '#INTRODUCE .detail_viewer',
        '#INTRODUCE [data-component-id]',
        '#INTRODUCE',
        '[data-name="INTRODUCE"][role="tabpanel"]',
        'xpath=//*[@id="INTRODUCE"]//div[contains(@data-component-id,"INTRODUCE")]//div[contains(@class,"se_component")]//div[last()]',
        'xpath=//*[@id="INTRODUCE"]/div/div[4]',
    ]

    for selector in content_selectors:
        element = page.query_selector(selector)
        if element is not None:
            print(f"Using content selector '{selector}' for {product_code}")
            return selector

    print("?곹뭹 ?곸꽭 而⑦뀗痢??곸뿭??李얠? 紐삵뻽?듬땲?? ?ㅻ깄???????None 諛섑솚")
    save_debug_snapshot(page, f"content_{product_code}")
    return None  # If the element is not found, return None


def crawl_page(page, df, seen_urls):
    time.sleep(1)
    page.wait_for_load_state("networkidle")  # ?섏씠吏 濡쒕뵫???꾨즺???뚭퉴吏 湲곕떎由쎈땲??
    products = find_elements(
        page,
        [
            "li:has(a[href*='/products/'])",
            "div:has(a[href*='/products/'])",
            "li[class*='flu7YgFW2k']",
        ],
    )
    if not products:
        print("?곹뭹 由ъ뒪????됲꽣媛 紐⑤몢 ?ㅽ뙣?덉뒿?덈떎. HTML ?ㅻ깄?룹쓣 ??ν빀?덈떎.")
        save_debug_snapshot(page, "product_list")
    duplicate_detected = False

    for i, product in enumerate(products):
        # if i >= 5:  # ??媛쒖쓽 ?곹뭹留??щ·留곹븯怨?猷⑦봽瑜?以묐떒?⑸땲??
        #     break

        product_data = get_product_data(page, product, i, len(products))

        # get_product_data媛 None??諛섑솚?섎㈃ ?대떦 ?쒗뭹??嫄대꼫?곷땲??
        if product_data is None:
            print(f"Skipping product at index {i} as get_product_data returned None.")
            continue

        product_url = product_data['Product_URL'][0]

        if product_url == "N/A" or not product_url:
            print("?곹뭹 URL 異붿텧 ?ㅽ뙣濡???ぉ??嫄대꼫?곷땲??")
            continue

        if product_url in seen_urls:
            print('Duplicate product detected: ', product_url)
            duplicate_detected = True
            break
        else:
            seen_urls.add(product_url)

        df = pd.concat([df, product_data], ignore_index=True)

        if MAX_PRODUCTS_PER_PAGE and (i + 1) >= MAX_PRODUCTS_PER_PAGE:
            print(f"Reached MAX_PRODUCTS_PER_PAGE={MAX_PRODUCTS_PER_PAGE}, stop crawling this page.")
            break

    return df, duplicate_detected


def extract_product_details(product):
    title_element = first_available(
        product,
        [
            "strong[aria-hidden='false']",
            "[data-testid='PRODUCT_CARD_TITLE']",
            "a[href*='/products/'] strong",
            "span[class*='ProductCard__Title']",
            "strong._26YxgX-Nu5",
        ],
    )
    if title_element:
        title = title_element.inner_text().strip()
    else:
        title = product.inner_text().splitlines()[0].strip() if product.inner_text() else "N/A"

    price_element = first_available(
        product,
        [
            "[data-testid='PRODUCT_CARD_PRICE']",
            "span:has-text('??)",
            "strong span:has-text('??)",
            "span._2DywKu0J_8",
        ],
    )
    if price_element:
        price = extract_price_from_text(price_element.inner_text())
    else:
        price = extract_price_from_text(product.inner_text())

    url_element = first_available(
        product,
        [
            "a[href*='/products/'][role='link']",
            "a[href*='/products/']",
            "a._2id8yXpK_k",
        ],
    )

    if url_element:
        raw_url = url_element.get_attribute("href")
        if raw_url and raw_url.startswith("/"):
            product_url = base_url + raw_url
        else:
            product_url = raw_url
    else:
        product_url = None

    product_code = product_url.split('/')[-1] if product_url else "N/A"
    return title, price, product_url or "N/A", product_code


def original_shipping_fee(page):
    print(f"Current page URL: {page.url}")  # ?꾩옱 ?섏씠吏 URL 異쒕젰

    shipping_selectors = [
        "xpath=//*[contains(@class,'delivery') and contains(text(),'??)]",
        "xpath=//span[contains(text(),'諛곗넚鍮?)]/following-sibling::*[1]",
        "xpath=//*[contains(text(),'諛곗넚鍮?) and contains(text(),'??)]",
        "xpath=//*[contains(text(),'諛섑뭹諛곗넚鍮?) and contains(text(),'??)]",
    ]

    for selector in shipping_selectors:
        element = page.query_selector(selector)
        if not element:
            continue
        element_text = element.inner_text().strip()
        if "臾대즺諛곗넚" in element_text:
            print("諛곗넚鍮? 臾대즺諛곗넚")
            return "0"
        digits = re.findall(r"[\d,]+", element_text)
        if digits:
            value = digits[0].replace(",", "")
            print(f"諛곗넚鍮? {value}")
            return value

    body_text = ""
    try:
        body_text = page.inner_text("body")
    except Exception:
        pass

    if "臾대즺諛곗넚" in body_text:
        print("諛곗넚鍮? 臾대즺諛곗넚(蹂몃Ц ?먯?)")
        return "0"

    for pattern in [r"諛곗넚鍮?s*[:竊??\s*([\d,]+)\s*??, r"諛섑뭹諛곗넚鍮?s*[:竊??\s*([\d,]+)\s*??]:
        match = re.search(pattern, body_text)
        if match:
            value = match.group(1).replace(",", "")
            print(f"諛곗넚鍮?蹂몃Ц ?먯?): {value}")
            return value

    print("Shipping fee element not found, ?????N/A 諛섑솚")
    save_debug_snapshot(page, "shipping_fee")
    return "N/A"


def option_crawl(page):
    option_data = {}

    option_triggers = page.query_selector_all('[data-shp-area$="optselect"]')
    if not option_triggers:
        option_triggers = page.query_selector_all(
            'a[role="button"][aria-haspopup="listbox"], button[aria-haspopup="listbox"]'
        )
        if option_triggers:
            print("Fallback option selector ?ъ슜 (listbox 踰꾪듉 湲곕컲)")

    option_index = 0
    while True:
        current_triggers = page.query_selector_all('[data-shp-area$="optselect"]')
        if not current_triggers:
            current_triggers = option_triggers

        if option_index >= len(current_triggers):
            break

        trigger = current_triggers[option_index]
        data_area = (trigger.get_attribute("data-shp-area") or "")
        if data_area and "optselect" not in data_area:
            option_index += 1
            continue

        option_index += 1
        category = trigger.get_attribute("aria-label") or trigger.inner_text().strip()
        if not category or category in {"?좏깮", ""}:
            category = f"?듭뀡{option_index}"

        try:
            trigger.click()
        except PlaywrightTimeoutError:
            print(f"{category} ?대┃ ?ㅽ뙣")
            continue

        try:
            dropdown = page.wait_for_selector("ul[role=\"listbox\"]", timeout=15000)
        except PlaywrightTimeoutError:
            print(f"{category} ?듭뀡 由ъ뒪??濡쒕뱶 ?ㅽ뙣")
            continue

        items = dropdown.query_selector_all("[role='option'], a, li")
        current_options = []
        current_prices = []
        seen_names = set()

        for item in items:
            option_text = item.inner_text().strip()
            if not option_text:
                continue
            price_match = re.search(r'\(([+\-]?[\d,]+)??)', option_text)
            if price_match:
                price_value = int(price_match.group(1).replace(',', ''))
                name = re.sub(r'\(([+\-]?[\d,]+)??)', '', option_text).strip()
            else:
                price_value = 0
                name = option_text

            # De-duplicate by normalized visible text to prevent double capture
            name_norm = re.sub(r"\s+", " ", name)
            if name_norm in seen_names:
                continue
            seen_names.add(name_norm)
            current_options.append(name_norm)
            current_prices.append(price_value)

        if not current_options:
            print(f"{category} ?듭뀡 ?뺣낫瑜?李얠? 紐삵뻽?듬땲??")
            continue

        option_data[category] = {
            '?섏쐞?듭뀡?쒕ぉ': current_options,
            '?섏쐞?듭뀡媛寃?: current_prices,
        }

        # ?꾩쓽 ?듭뀡???좏깮???ㅼ쓬 ?④퀎 吏꾪뻾
        selectable = []
        for item in items:
            try:
                disabled = item.evaluate("node => node.getAttribute('aria-disabled') === 'true'")
            except Exception:
                disabled = False
            if not disabled:
                selectable.append(item)

        if selectable:
            random.choice(selectable).click()
            time.sleep(0.5)

    return option_data


def image_crawl(page):
    main_candidates = find_elements(
        page,
        [
            "img[alt='??쒖씠誘몄?']",
            "img[alt*='???][src*='shop-phinf']",
            "div[id='content'] img[src*='shop-phinf']",
        ],
    )
    thumbnail_elements = find_elements(
        page,
        [
            "img[alt^='異붽??대?吏']",
            "button[aria-label^='?몃꽕??] img",
            "ul[class*='thumbnail'] img",
        ],
    )

    if not thumbnail_elements:
        thumbnail_elements = page.query_selector_all("img[src*='shop-phinf']")

    image_elements = []
    if main_candidates:
        image_elements.extend(main_candidates)
    if thumbnail_elements:
        image_elements.extend(thumbnail_elements)

    if not image_elements:
        try:
            fallback = page.wait_for_selector(
                'xpath=//*[@id="content"]//img[contains(@src,"shop-phinf")]',
                timeout=5000,
            )
            if fallback:
                image_elements = [fallback]
        except Exception:
            pass

    if not image_elements:
        print("No images found on the page.")
        return [], []

    thumbnail_urls = []
    for element in image_elements:
        src = element.get_attribute("src")
        if not src:
            continue
        thumbnail_urls.append(src.split("?")[0])

    # Deduplicate while preserving order
    seen = set()
    unique_urls = []
    for url in thumbnail_urls:
        if url in seen:
            continue
        seen.add(url)
        unique_urls.append(url)

    if not unique_urls:
        print("Image URLs could not be extracted.")
        return [], []

    def url_prefix(u):
        filename = u.split("/")[-1]
        digits = "".join(ch for ch in filename if ch.isdigit())
        return digits[:3] if digits else filename[:3]

    prefixes = [url_prefix(url) for url in unique_urls]
    counts = Counter(prefixes)
    if not counts:
        return unique_urls, []

    most_common = counts.most_common(1)[0][0]
    common_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix == most_common]
    different_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix != most_common]

    return common_urls, different_urls





def content_crawl(page, product_code, element_selector):
    time.sleep(1)
    page.wait_for_load_state("load")

    # element_selector媛 ?좏슚?쒖? ?뺤씤
    if not element_selector:
        print("Invalid element selector. Moving to the next item.")
        return None

    element = page.query_selector(element_selector)
    if element is None:
        # element_selector???좏슚?섏?留? ?대떦 ?붿냼瑜?李얠? 紐삵븳 寃쎌슦
        print("No valid element found for the selector. Moving to the next item.")
        return None

    content = element.inner_html()
    soup = BeautifulSoup(content, 'html.parser')

    if '怨꾩냽?⑸땲?? in soup.get_text():
        print("found 怨꾩냽?⑸땲??)
        return None

    css_link = soup.new_tag("link", rel="stylesheet",
                            href="https://static-resource-smartstore.pstatic.net/smartstore/p/static/20230630180923/common.css")
    if soup.head:
        soup.head.append(css_link)
    else:
        head_tag = soup.new_tag("head")
        head_tag.append(css_link)
        soup.insert(0, head_tag)

    for button in soup.find_all('button'):
        button.decompose()

    for img in soup.find_all('img', attrs={'data-src': True}):
        img['src'] = img['data-src']
        del img['data-src']

    target_url = "https://rapid-up.s3.ap-northeast-2.amazonaws.com/dev/gray-line.png"
    new_url = "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/gray-line.png"
    for img in soup.find_all('img', src=target_url):
        img['src'] = new_url

    text_to_remove = '* {text-align: center;}  #mycontents11 img{max-width: 100%;}'
    if text_to_remove in soup.get_text():
        soup = BeautifulSoup(str(soup).replace(text_to_remove, ''), 'html.parser')

    disallowed_attrs = ['area-hidden', 'data-linkdata', 'data-linktype', 'onclick', 'style', 'class']
    for attr in disallowed_attrs:
        for tag in soup.find_all(attrs={attr: True}):
            del tag[attr]

    for a in soup.find_all('a', attrs={'data-linkdata': True}):
        img = a.find('img')
        if img:
            src = img.get('src', '')
            data_src = img.get('data-src', '')
            if not src:
                src = data_src
                img['src'] = src
                del img['data-src']
            # #?뱀젙 ?쒓렇 ?덉뿉 荑좊? ?대?吏 ?쒓굅
            # if src.startswith('https://rapid-up.s3.ap-northeast-2.amazonaws.com'):
            #     img.decompose()
            # elif not src.startswith(('https://cdn.011st.com', 'https://img.alicdn.com')):
            #     new_src = upload_to_oracle_cloud(src)
            #     img['src'] = new_src
    #紐⑤뱺 荑좊? ?대?吏 ?쒓굅
    for img in soup.find_all('img'):
        src = img.get('src', '')
        if src.startswith(('https://rapid-up.s3.ap-northeast-2.amazonaws.com', 'https://cdn.heyseller.kr', 'https://ai.esmplus.com/')):
            img.decompose()

    # img_tags = soup.find_all('img')
    # print(f"Number of images before: {len(img_tags)}")
    #
    # # ?쒓굅?섎뒗 ?대?吏??src ?띿꽦 ?뺤씤
    # img_tags = soup.find_all('img')
    # print(f"Number of images before: {len(img_tags)}")
    #
    # # ?쒓굅?섎뒗 ?대?吏??src ?띿꽦 ?뺤씤
    # if len(img_tags) > 0:
    #     print(f"First image src before removal: {img_tags[0].get('src', 'No src attribute')}")
    # if len(img_tags) > 1:
    #     print(f"Last image src before removal: {img_tags[-1].get('src', 'No src attribute')}")

    # # 泥?踰덉㎏? 留덉?留??대?吏 ?쒓굅
    # if len(img_tags) > 3:
    #     img_tags[0].decompose()
    #     img_tags = soup.find_all('img')  # 由ъ뒪???낅뜲?댄듃
    #     img_tags[-1].decompose()
    # elif len(img_tags) > 1:
    #     img_tags[0].decompose()
    #     img_tags = soup.find_all('img')  # 由ъ뒪???낅뜲?댄듃
    #     img_tags[-1].decompose()

    # # 'heyseller'瑜??ы븿?섎뒗 ?대?吏 ?쒓굅
    # for img in img_tags[:]:  # 蹂듭궗蹂몄쓣 ?쒗쉶?섏뿬 ?먮낯 由ъ뒪?몃? ?섏젙
    #     if img and img.attrs and 'heyseller' in img.get('src', ''):
    #         img.decompose()

    img_tags = soup.find_all('img')  # 理쒖쥌 ?대?吏 ?쒓렇 由ъ뒪???낅뜲?댄듃
    print(f"Number of images after all removals: {len(img_tags)}")

    soup = insert_and_remove_images(soup)

    for img_tag in soup.find_all('img'):
        img_tag['style'] = 'display: block; margin-left: auto; margin-right: auto; margin-bottom: 10px;'

    for h1 in soup.find_all('h1'):
        h1['style'] = 'text-align: center; font-size: 30px; margin-bottom: 20px;'

    text_elements = ['p', 'div', 'span', 'li', 'a']
    for tag_name in text_elements:
        for element in soup.find_all(tag_name):
            existing_style = element.get('style', '')
            new_style = f"{existing_style}; text-align: center; font-size: 18px; margin-bottom: 30px;"
            element['style'] = new_style.strip()

    return pd.DataFrame({'Content': [str(soup)]})




def insert_and_remove_images(soup):
    img_srcs_to_insert = ['https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/top.png',
                          'https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/bottom.png',
                          'https://coudae.s3.ap-northeast-2.amazonaws.com/A00412936/cloud/7290.png']

    img_tag_top = soup.new_tag('img', src=img_srcs_to_insert[0],
                               style="display: block; margin-left: auto; margin-right: auto;")
    img_tag_middle = soup.new_tag('img', src=img_srcs_to_insert[2],
                                  style="display: block; margin-left: auto; margin-right: auto;")
    img_tag_bottom = soup.new_tag('img', src=img_srcs_to_insert[1],
                                  style="display: block; margin-left: auto; margin-right: auto;")

    first_tag = next(soup.children)
    last_tag = next(reversed(soup.contents))

    first_tag.insert_before(img_tag_top)
    last_tag.insert_after(img_tag_bottom)

    img_srcs_to_remove = ['', '']

    for img_src in img_srcs_to_remove:
        img_to_remove = soup.find_all('img', attrs={'src': img_src})
        for img in img_to_remove:
            img.decompose()

    return soup


def return_shipping_fee(total_price):
    fee = total_price * 0.25
    if fee > 200000:
        fee = 200000
    return fee



def title_edit(title):
    # 以묐났 ?⑥뼱 ?쒓굅
    title_split = title.split(' ')
    title_split = list(dict.fromkeys(title_split))  # 以묐났 ?쒓굅

    # 留덉?留??⑥뼱? 留덉?留됱뿉????踰덉㎏ ?⑥뼱 援먯껜
    if len(title_split) >= 2:
        title_split[-1], title_split[-2] = title_split[-2], title_split[-1]

    # ?ㅼ떆 臾몄옄?대줈 議고빀
    title = ' '.join(title_split)
    return title


def get_product_data(page, product, i, num_products):
    title, price, product_url, product_code = extract_product_details(product)
    
    # 臾몄옄??泥섎━ 諛⑹떇 ?곸슜
    title = title.replace('\xa0', ' ')
    
    
    print(f"Product {i + 1}/{num_products}: {title}, {price} won, {product_url}")

    product_page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(product_page)
    product_page.goto(product_url)
    product_page.wait_for_load_state("load")
    try:
        product_page.wait_for_load_state("networkidle", timeout=10000)
    except PlaywrightTimeoutError:
        pass

    scripts = product_page.query_selector_all('script')

    category_df = pd.read_excel(NAVER_CATEGORY_PATH, header=None)
    small_category_dict = pd.Series(category_df[0].values, index=category_df[3]).to_dict()
    tiny_category_dict = pd.Series(category_df[0].values, index=category_df[4]).to_dict()

    category = None
    for script in scripts:
        script_content = script.inner_text()
        if "category" in script_content:
            json_data = json.loads(script_content)
            if 'category' in json_data:
                category = json_data['category']
                break

    if category is not None:
        print(f"Category: {category}")

        # Parse the category text into a list
        category_list = category.split(">")
        large_category = category_list[0].strip()
        medium_category = category_list[1].strip()
        small_category = category_list[2].strip() if len(category_list) > 2 else None
        tiny_category = category_list[3].strip() if len(category_list) > 3 else None
    else:
        # 'category'媛 None??寃쎌슦, 紐⑤뱺 移댄뀒怨좊━ 愿??蹂?섎뱾??None?쇰줈 ?ㅼ젙
        category_list = []
        large_category = None
        medium_category = None
        small_category = None
        tiny_category = None

    # Determine the smallest category and its type
    if tiny_category is not None:
        smallest_category = tiny_category
        smallest_category_type = 'tiny'
        naver_category_number = tiny_category_dict.get(tiny_category)
    elif small_category is not None:
        smallest_category = small_category
        smallest_category_type = 'small'
        naver_category_number = small_category_dict.get(small_category)
    else:
        # tiny_category 諛?small_category媛 紐⑤몢 None??寃쎌슦
        smallest_category = None
        smallest_category_type = None
        naver_category_number = None

    # Print out the category hierarchy and the smallest category
    print(
        f"Smallest Category('{smallest_category_type}') : {smallest_category}, Naver category number: {naver_category_number}")

    options = option_crawl(product_page)
    print("Options:", options)

    common_urls, different_urls = image_crawl(product_page)
    print(f"common_urls: {common_urls}")

    main_image = None
    other_images = []

    if common_urls:
        main_image = common_urls[0].replace('?type=m510', '')
        other_images = [url.replace('?type=m510', '') for url in common_urls[1:]]
        print(f"other_images: {other_images}")  # other_images 異쒕젰
    else:
        print("No common images found")
        try:
            image_element = page.wait_for_selector('xpath=//*[@id="content"]/div/div[2]/div[1]/div[1]/div[1]/img', timeout=2000)
            image_url = image_element.get_attribute("src")
            main_image = image_url.replace('?type=m510', '')
        except:
            print("No main image found")

    print("Main image:", main_image)
    print("Other images:", other_images)

    # Check if main_image is not None before uploading
    #if main_image:
    #    main_image = upload_to_oracle_cloud(main_image)
    #else:
    #    print("Main image is None, skipping upload to Oracle Cloud.")

    # Ensure only valid URLs are uploaded for other_images
    #other_images = [upload_to_oracle_cloud(url) for url in other_images if url]

    print("URLs not starting with most common three digits:")
    for url in different_urls:
        print(url)

    element_selector = find_content_element(product_page, product_code)
    content = content_crawl(product_page, product_code, element_selector)
    shipping_fee = original_shipping_fee(product_page)

    def to_int(value):
        if value in (None, "N/A"):
            return 0
        digits = re.sub(r"[^\d]", "", str(value))
        return int(digits) if digits else 0

    shipping_fee_int = to_int(shipping_fee)
    price_int = to_int(price)
    total_price = price_int + shipping_fee_int

    # Check if the returned content is a string
    if isinstance(content, str):
        content_df = pd.DataFrame({'Content': [content]})
    else:
        content_df = content

    product_df = pd.DataFrame({
        'Product': [title],
        'Price': [price],
        'Shipping_Fee': [shipping_fee_int],  # Updated to integer value
        'Total_Price': [total_price],  # New column
        'Main_Image': [main_image],
        'Other_Images': [other_images],
        'Options': [options],
        'Product_URL': [product_url],
        'Naver_Category_Number': [naver_category_number]
    })

    product_df = pd.concat([product_df, content_df], axis=1)
    product_page.close()

    return product_df



def write_to_excel(df, excel_path, seen_urls):
    from openpyxl import load_workbook
    import pandas as pd

    book = load_workbook(excel_path)
    sheet = book['?쇨큵?깅줉']

    b_start_row = c_start_row = e_start_row = h_start_row = ad_start_row = r_start_row = s_start_row = t_start_row = i_start_row = 3
    ap_start_row = aq_start_row = 3

    for i, item in enumerate(df['Naver_Category_Number'], start=b_start_row):
        sheet['B' + str(i)] = item
    for j, item in enumerate(df['Product'], start=c_start_row):
        sheet['C' + str(j)] = item
    for k, (product_price, shipping_fee) in enumerate(zip(df['Price'], df['Shipping_Fee']), start=e_start_row):
        total_price = float(product_price.replace(',', '')) + shipping_fee
        selling_price = total_price - 0.01 * total_price
        selling_price_rounded = round(selling_price / 100.0) * 100.0
        sheet['E' + str(k)] = selling_price_rounded

    for l, item in enumerate(df.iterrows(), start=h_start_row):
        sheet['H' + str(l)] = "議고빀??
    for m in range(ad_start_row, ad_start_row + len(df)):
        selling_price_rounded = sheet['E' + str(m)].value




###########諛붿젮留덉폆
        if selling_price_rounded <= 20000:
            ad_value = 2903608
        elif 20001 <= selling_price_rounded <= 30000:
            ad_value = 2904260
        elif 30001 <= selling_price_rounded <= 40000:
            ad_value = 2904261
        elif 40001 <= selling_price_rounded <= 60000:
            ad_value = 2904262
        elif 60001 <= selling_price_rounded <= 80000:
            ad_value = 2904268
        elif 80001 <= selling_price_rounded <= 100000:
            ad_value = 2904272
        elif 100001 <= selling_price_rounded <= 150000:
            ad_value = 2904276
        elif 150001 <= selling_price_rounded <= 400000:
            ad_value = 2904278
        elif 400001 <= selling_price_rounded <= 600000:
            ad_value = 2904279
        elif 600001 <= selling_price_rounded <= 1000000:
            ad_value = 2904281
        elif 1000001 <= selling_price_rounded <= 9999999:
            ad_value = 2904284


        sheet['AD' + str(m)] = ad_value


    for row, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['U' + str(row)] = "?곸꽭?섏씠吏 李몄“"
        sheet['V' + str(row)] = "?곸꽭?섏씠吏 李몄“"
        sheet['Y' + str(row)] = "0200037"
        sheet['Z' + str(row)] = "援щℓ???
        sheet['AZ' + str(row)] = "010-3973-3119"
        sheet['BA' + str(row)] = "蹂몃Ц ?덈궡臾?李몄“"

    for l, item in enumerate(df['Options'], start=r_start_row):
        option_titles = []
        option_prices = []
        option_categories = []
        for index, key in enumerate(item):
            option_categories.append(key)
            if '?섏쐞?듭뀡?쒕ぉ' in item[key]:
                option_titles.append(', '.join(item[key]['?섏쐞?듭뀡?쒕ぉ']))
            if '?섏쐞?듭뀡媛寃? in item[key]:
                option_prices.append(', '.join(map(str, item[key]['?섏쐞?듭뀡媛寃?])))
        sheet['I' + str(l)] = '\n'.join(option_categories)
        sheet['J' + str(l)] = '\n'.join(option_titles)
        sheet['K' + str(l)] = '\n'.join(option_prices)

    # 'K' ??媛믪쓣 湲곕컲?쇰줈 'L' ??媛믪쓣 ?낅뜲?댄듃?섎뒗 遺遺?
    if 'Options' in df.columns:
        for row in range(h_start_row, h_start_row + len(df)):
            item_options = df.at[row - h_start_row, 'Options']
            option_prices = []
            for option in item_options.values():
                if '?섏쐞?듭뀡媛寃? in option:
                    option_prices.extend(option['?섏쐞?듭뀡媛寃?])

            if option_prices:
                num_prices = len(option_prices)
                l_values = ', '.join(['99'] * num_prices)
                sheet['L' + str(row)] = l_values
            else:
                # ?듭뀡 媛寃⑹씠 ?녿뒗 寃쎌슦, 'L' ?댁쓣 鍮꾩썙 ?먭굅??湲곕낯媛믪쓣 ?ㅼ젙
                sheet['L' + str(row)] = "99"  # ?먮뒗 "湲곕낯媛?
    else:
        for row in range(h_start_row, h_start_row + len(df)):
            # 'Options' ?댁씠 ?녿뒗 寃쎌슦, 'L' ?댁쓣 鍮꾩썙 ?먭굅??湲곕낯媛믪쓣 ?ㅼ젙
            sheet['L' + str(row)] = "99"  # ?먮뒗 "湲곕낯媛?



    for n, item in enumerate(df['Main_Image'], start=r_start_row):
        sheet['R' + str(n)] = item
    for o, item in enumerate(df['Other_Images'], start=s_start_row):
        if item is not None:
            sheet['S' + str(o)] = "\n".join(str(img) for img in item if img is not None)
        else:
            sheet['S' + str(o)] = ""

    # 'Content' ?댁쓽 議댁옱 ?щ?瑜??뺤씤?섍퀬 泥섎━
    if 'Content' in df.columns:
        for p, item in enumerate(df['Content'], start=t_start_row):
            if isinstance(item, float):
                item = str(item)
            print(f"Row {p}, Content: {item[:100]}")

            # ?곗씠???묒? ?뚯씪??湲곕줉
            sheet['T' + str(p)] = item
    else:
        print("No 'Content' column found in DataFrame")

    for q, url in enumerate(df['Product_URL'], start=i_start_row):
        product_code = url.split('/')[-1]
        try:
            selling_code = str(int(product_code) * 2)
        except ValueError:
            selling_code = str(random.randint(10000000, 99999999)) + 'R'
        sheet['A' + str(q)] = selling_code
    for r, total_price in enumerate(df['Total_Price'], start=ap_start_row):
        return_fee = return_shipping_fee(total_price)
        return_fee_rounded = round(return_fee / 100.0) * 100
        sheet['AP' + str(r)] = return_fee_rounded
        sheet['AQ' + str(r)] = return_fee_rounded * 2  # Set 'AQ' column to double of return_fee_rounded

    book.save(excel_path)


    # Here we load the workbook again
    book = load_workbook(excel_path)
    sheet = book['?쇨큵?깅줉']

    for i in range(3, sheet.max_row + 1):
        if not sheet['A' + str(i)].value:
            sheet.delete_rows(i, sheet.max_row - i + 1)
            break

    book.save(excel_path)
    if os.name == "nt":
        os.system(f'start \"\" \"excel.exe\" \"{excel_path}\"')
    else:
        print(f"Excel file saved to {excel_path}. Automatic Excel launch is skipped on non-Windows platforms.")



def write_to_excel2(df, excel_path2):
    df2 = pd.DataFrame({
        'Product_URL': df['Product_URL'],
        'Numbering': range(1, len(df) + 1),
        'Product_Title': df['Product'],
        'Product_Price': df['Price'],
        'Shipping_Fee': df['Shipping_Fee']  # 異붽???諛곗넚鍮???ぉ
    })
    with pd.ExcelWriter(excel_path2) as writer:
        df2.to_excel(writer, index=False)



# userAgentStrings = [
#     # 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.2227.0 Safari/537.36',
#     # 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
#     # 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.3497.92 Safari/537.36',
#     'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36 Edg/92.0.902.78',
# ]

maybe_launch_chrome_devtools()

if CRAWLER_DRY_RUN:
    print("CRAWLER_DRY_RUN=1 ?뚮옒洹몃줈 ?명빐 Playwright ?щ·留?蹂몃룞?묒쓣 ?앸왂?⑸땲??")
    sys.stdout = original
    f.close()
    sys.exit(0)

with sync_playwright() as p:
    browser_name = os.getenv("PLAYWRIGHT_BROWSER", "chromium").lower()
    if browser_name not in {"chromium", "firefox", "webkit"}:
        browser_name = "chromium"

    connect_url_env = os.getenv("PLAYWRIGHT_CONNECT_URL") if (browser_name == "chromium" and not FORCE_LOCAL_PLAYWRIGHT) else None
    connect_url_source = None
    connect_url = None
    if connect_url_env and connect_url_env.strip():
        sanitized_url = normalize_cdp_url(connect_url_env)
        if sanitized_url and sanitized_url != connect_url_env.strip():
            print(
                "PLAYWRIGHT_CONNECT_URL 媛믪쓣 HTTP(S) CDP 猷⑦듃 URL濡??뺢퇋?뷀뻽?듬땲??\n"
                f"?낅젰媛? {connect_url_env}\n"
                f"?곸슜媛? {sanitized_url}"
            )
        connect_url = sanitized_url
        connect_url_source = "env"
    elif browser_name == "chromium":
        connect_url = DEFAULT_CONNECT_URL
        connect_url_source = "default"

    headless_mode = os.getenv("PLAYWRIGHT_HEADLESS", "1").lower() not in {"false", "0", "no"}
    should_close_browser = connect_url is None

    if browser_name == "chromium":
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-features=NetworkService",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-web-security",
            "--disable-dev-shm-usage",
            "--proxy-bypass-list=<-loopback>",
            "--disable-accelerated-2d-canvas",
            "--disable-gpu",
            "--no-zygote",
            # '--single-process' ?듭뀡? ?쇰? WSL/而⑦뀒?대꼫 ?섍꼍?먯꽌 sandbox_host ?ㅻ쪟瑜??좊컻?섎?濡??쒖쇅?쒕떎.
        ]

        if FORCE_LOCAL_PLAYWRIGHT:
            connect_url = None  # 媛뺤젣 濡쒖뺄 ?ㅽ뻾
            should_close_browser = True
            print("FORCE_LOCAL_PLAYWRIGHT=1: CDP ?곌껐???앸왂?섍퀬 濡쒖뺄 Chromium???ㅽ뻾?⑸땲??")

        if connect_url:
            # 1李??곌껐 ?쒕룄: ?섍꼍蹂???먮뒗 湲곕낯媛?
            ok = verify_cdp_endpoint(connect_url)
            original_url = connect_url
            if not ok and running_on_wsl():
                # WSL?먯꽌 127.0.0.1/localhost瑜??ъ슜 以묒씠硫?Windows ?몄뒪??IP濡??먮룞 ?泥??ъ떆??
                parts = urlsplit(connect_url)
                host = parts.hostname or "127.0.0.1"
                port = parts.port or DEFAULT_CDP_PORT
                if host in {"127.0.0.1", "localhost"}:
                    alt_host = detect_windows_host_from_wsl()
                    if alt_host:
                        alt_url = urlunsplit((parts.scheme, f"{alt_host}:{port}", "", "", ""))
                        print(
                            "WSL ?섍꼍?먯꽌 127.0.0.1 ???Windows ?몄뒪??IP濡??ъ떆?꾪빀?덈떎.\n"
                            f"??? {alt_url}"
                        )
                        if verify_cdp_endpoint(alt_url):
                            connect_url = alt_url
                            ok = True
                            connect_url_source = (connect_url_source or "env") + " (WSL fallback)"
                        else:
                            # 理쒗썑???섎떒: Windows curl???듯빐 ws endpoint 議고쉶 ??援먯껜 ?쒕룄
                            cmd_parts = locate_cmd_invocation()
                            if cmd_parts is not None:
                                try:
                                    out = subprocess.run(
                                        [*cmd_parts, f"curl -sS http://127.0.0.1:{port}/json/version"],
                                        check=True,
                                        capture_output=True,
                                        text=True,
                                        timeout=8,
                                    )
                                    raw = out.stdout
                                    json_str = _extract_first_json_object(raw)
                                    if json_str:
                                        data = json.loads(json_str)
                                        ws = data.get("webSocketDebuggerUrl")
                                        if isinstance(ws, str) and ws:
                                            ws_parts = urlsplit(ws)
                                            # localhost瑜??ㅼ젣 ?몄뒪?몃줈 移섑솚
                                            ws_url = urlunsplit((
                                                "ws" if ws_parts.scheme.startswith("ws") else "ws",
                                                f"{alt_host}:{port}",
                                                ws_parts.path,
                                                ws_parts.query,
                                                ws_parts.fragment,
                                            ))
                                            connect_url = ws_url
                                            ok = True
                                            connect_url_source = (connect_url_source or "env") + " (WSL ws fallback)"
                                except Exception:
                                    pass
            if not ok and running_on_wsl():
                # 異붽? ?쒕룄: ?뱀떆 Windows媛 ?꾨땶 WSL ?대? Chrome???좎엳??寃쎌슦 127.0.0.1濡쒕룄 ?ㅼ떆 ?쒕룄
                parts = urlsplit(original_url)
                port = parts.port or DEFAULT_CDP_PORT
                local_url = urlunsplit((parts.scheme, f"127.0.0.1:{port}", "", "", ""))
                if local_url != original_url:
                    print(f"異붽? 濡쒖뺄 ?ъ떆?? {local_url}")
                    if verify_cdp_endpoint(local_url):
                        connect_url = local_url
                        ok = True
                        connect_url_source = (connect_url_source or "env") + " (local retry)"

            if not ok and running_on_wsl():
                # 理쒖쥌 WSL ?꾩슜 ws ?붾뱶?ъ씤???泥??쒕룄 (?몄뒪??醫낅쪟 臾닿?)
                parts_any = urlsplit(original_url)
                port_any = parts_any.port or DEFAULT_CDP_PORT
                alt_host_any = detect_windows_host_from_wsl() or (parts_any.hostname or DEFAULT_CDP_HOST)
                cmd_parts = locate_cmd_invocation()
                if cmd_parts is not None:
                    try:
                        out = subprocess.run(
                            [*cmd_parts, f"curl -sS http://127.0.0.1:{port_any}/json/version"],
                            check=True,
                            capture_output=True,
                            text=True,
                            timeout=8,
                        )
                        raw = out.stdout
                        json_str = _extract_first_json_object(raw)
                        if json_str:
                            data = json.loads(json_str)
                            ws = data.get("webSocketDebuggerUrl")
                            if isinstance(ws, str) and ws:
                                ws_parts = urlsplit(ws)
                                ws_url = urlunsplit((
                                    "ws" if ws_parts.scheme.startswith("ws") else "ws",
                                    f"{alt_host_any}:{port_any}",
                                    ws_parts.path,
                                    ws_parts.query,
                                    ws_parts.fragment,
                                ))
                                connect_url = ws_url
                                ok = True
                                connect_url_source = (connect_url_source or "default") + " (WSL ws fallback-any)"
                    except Exception:
                        pass

            if not ok:
                if REQUIRE_CDP_CONNECTION:
                    print(f"CDP ?곌껐 ?뺤씤???ㅽ뙣?덉뒿?덈떎. URL: {original_url}")
                    sys.exit(1)
                print(
                    "CDP ?곌껐 ?뺤씤???ㅽ뙣?섏뿬 濡쒖뺄 Chromium ?ㅽ뻾?쇰줈 ?泥댄빀?덈떎. "
                    "Chrome DevTools 諛곗튂 ?ㅽ뻾 ?곹깭? ?ы듃/二쇱냼瑜??뺤씤?섏꽭??"
                )
                connect_url = None
                connect_url_source = None
            else:
                if connect_url_source == "env":
                    print(f"?섍꼍蹂??PLAYWRIGHT_CONNECT_URL ({connect_url})???곌껐?⑸땲??")
                elif connect_url_source == "default":
                    print(f"湲곕낯 Chrome DevTools ?붾뱶?ъ씤??{connect_url})???곌껐?⑸땲??")

        should_close_browser = connect_url is None

        if connect_url:
            browser = p.chromium.connect_over_cdp(connect_url)
            if browser.contexts:
                context = browser.contexts[0]
            else:
                context = browser.new_context()
        else:
            browser = p.chromium.launch(
                headless=headless_mode, args=launch_args, chromium_sandbox=False
            )
            context = browser.new_context()
    elif browser_name == "firefox":
        browser = p.firefox.launch(headless=headless_mode)
        context = browser.new_context()
    else:
        browser = p.webkit.launch(headless=headless_mode)
        context = browser.new_context()

    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(context)
    df = pd.DataFrame(columns=['Product', 'Price', 'Product_URL'])
    script_dir = SCRIPT_DIR
    output_folder = script_dir / 'output'
    read_excel_path = output_folder / 'ExcelSaveTemplate_230109.xlsx'
    seen_urls = set()
    product_list_crawl(context, df, read_excel_path, seen_urls)
    try:
        context.close()
    finally:
        if should_close_browser:
            browser.close()

sys.stdout = original
f.close() 
