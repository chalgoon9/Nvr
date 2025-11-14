from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
try:
    from playwright_stealth import Stealth
except ImportError:
    Stealth = None
from collections import Counter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd
import random
import time
import shutil
import re
import os
import json
import requests
import sys
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode


# 페이지당 최대 크롤링 상품 수 (테스트 기본값 5개)
DEFAULT_MAX_PRODUCTS_PER_PAGE = 5
MAX_PRODUCTS_PER_PAGE = int(
    os.getenv("MAX_PRODUCTS_PER_PAGE", str(DEFAULT_MAX_PRODUCTS_PER_PAGE)) or DEFAULT_MAX_PRODUCTS_PER_PAGE
)

# 전체 실행에서 최대 수집 상품 수 (테스트 기본값 5개)
DEFAULT_MAX_PRODUCTS_TOTAL = DEFAULT_MAX_PRODUCTS_PER_PAGE
MAX_PRODUCTS_TOTAL = int(
    os.getenv("MAX_PRODUCTS_TOTAL", str(DEFAULT_MAX_PRODUCTS_TOTAL)) or DEFAULT_MAX_PRODUCTS_TOTAL
)


# .env 로더 (python-dotenv 미설치 시 최소 파서)
try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None


def fallback_load_dotenv(dotenv_path):
    path = Path(dotenv_path)
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        return False
    except OSError as exc:
        print(f".env 파일을 읽는 중 오류가 발생했습니다: {exc}")
        return False

    loaded = False
    saw_assignments = False
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        saw_assignments = True
        key = key.strip()
        value = value.strip().strip("\"'")
        if not key:
            continue
        if key not in os.environ:
            os.environ[key] = value
            loaded = True
    return loaded or saw_assignments


SCRIPT_DIR = Path(__file__).resolve().parent
LOG_FILE = SCRIPT_DIR / "log.txt"


def resolve_category_path() -> Path:
    # 우선 로컬 경로, 없으면 상위 디렉터리(AGENTS.md 가이드에 맞춤)
    local = SCRIPT_DIR / "naver_category.xlsx"
    parent = SCRIPT_DIR.parent / "naver_category.xlsx"
    if local.exists():
        return local
    if parent.exists():
        return parent
    # 마지막으로 로컬 경로를 반환(실패 시 런타임에서 에러 메시지 제공)
    return local


NAVER_CATEGORY_PATH = resolve_category_path()
CRAWLER_DRY_RUN = os.getenv("CRAWLER_DRY_RUN", "0").lower() in {"1", "true", "yes"}
WRAP_CONTENT_HTML = os.getenv("WRAP_CONTENT_HTML", "0").lower() in {"1", "true", "yes"}
DUMP_CONTENT_HTML = os.getenv("DUMP_CONTENT_HTML", "0").lower() in {"1", "true", "yes"}
DUMP_CONTENT_DIR = SCRIPT_DIR / "debug" / "content_outputs"

# 선택 페이지만 크롤링하는 디버그용 옵션(예: CRAWL_ONLY_PAGES="51,59").
# 지정되지 않으면 기존 범위(global_start_page~global_last_page) 전체를 처리합니다.
_only_pages_raw = os.getenv("CRAWL_ONLY_PAGES", "").strip()
CRAWL_ONLY_PAGES = None
if _only_pages_raw:
    try:
        parts = re.split(r"[\s,;]+", _only_pages_raw)
        CRAWL_ONLY_PAGES = [int(p) for p in parts if p]
    except Exception as exc:
        print(f"CRAWL_ONLY_PAGES 파싱 실패({_only_pages_raw}): {exc}")
        CRAWL_ONLY_PAGES = None

# 페이지 번호 이동 시 URL 쿼리 파라미터로 강제 점프 시도 여부
# 기본값은 비활성화(네이버는 URL 파라미터만으로 DOM이 바뀌지 않는 경우가 많음)
PAGE_JUMP_BY_QUERY = os.getenv("PAGE_JUMP_BY_QUERY", "0").lower() in {"1", "true", "yes"}

# 페이지 이동 과정 스크린샷 저장(디버깅 용도)
# 기본값 비활성화: 요청에 따라 캡처 중단
PAGINATION_DEBUG_SHOTS = os.getenv("PAGINATION_DEBUG_SHOTS", "0").lower() in {"1", "true", "yes"}

# 페이지네이션 전략: auto(기본) | next_only('다음'만 반복)
PAGINATION_STRATEGY = os.getenv("PAGINATION_STRATEGY", "auto").lower().strip()

def debug_shot(page, label):
    if not PAGINATION_DEBUG_SHOTS:
        return
    try:
        ts = time.strftime("%Y%m%d_%H%M%S")
        dbg = (SCRIPT_DIR / "debug")
        dbg.mkdir(exist_ok=True)
        path = dbg / f"pagination_{ts}_{label}.png"
        page.screenshot(path=str(path), full_page=True)
        print(f"Saved screenshot: {path}")
    except Exception as exc:
        print(f"Failed to take screenshot({label}): {exc}")

STEALTH_HELPER = Stealth() if Stealth is not None else None
if STEALTH_HELPER is None:
    print(
        "playwright_stealth 모듈에서 Stealth 클래스를 불러오지 못했습니다. "
        "탐지 회피 스크립트가 적용되지 않으니 chromium 환경에서는 추가 점검이 필요합니다."
    )


if load_dotenv:
    load_dotenv()
else:
    _ = fallback_load_dotenv(SCRIPT_DIR / ".env")


def first_available(node, selectors):
    for selector in selectors:
        try:
            element = node.query_selector(selector)
        except PlaywrightTimeoutError:
            continue
        if element:
            return element
    return None


def find_elements(page, selectors):
    for selector in selectors:
        try:
            elements = page.query_selector_all(selector)
        except PlaywrightTimeoutError:
            continue
        if elements:
            print(f"Selector '{selector}' matched {len(elements)} elements.")
            return elements
    return []


def save_debug_snapshot(page, prefix):
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    path = (SCRIPT_DIR / "debug")
    path.mkdir(exist_ok=True)
    out = path / f"{prefix}_{timestamp}.html"
    try:
        out.write_text(page.content(), encoding="utf-8")
        print(f"Saved debug snapshot: {out}")
    except Exception as exc:
        print(f"Failed to save debug snapshot: {exc}")


def save_debug_html(product_code, html_text, suffix):
    path = (SCRIPT_DIR / "debug")
    path.mkdir(exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = path / f"{product_code}_{suffix}_{timestamp}.html"
    try:
        filename.write_text(html_text, encoding="utf-8")
        print(f"[CONTENT][{product_code}] Saved debug HTML: {filename}")
    except Exception as exc:
        print(f"[CONTENT][{product_code}] Failed to save debug HTML: {exc}")


def extract_price_from_text(raw_text):
    match = re.search(r"([\d,]+)\s*원", raw_text)
    if match:
        return match.group(1).replace("\u200b", "").strip()
    digits = re.sub(r"[^\d]", "", raw_text)
    if not digits:
        return "N/A"
    try:
        return "{:,}".format(int(digits))
    except ValueError:
        return "N/A"


def has_numeric_chars(value):
    if value is None:
        return False
    return bool(re.search(r"\d", str(value)))


def normalize_price_value(value):
    if value in (None, "", "N/A"):
        return None
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return int(value)
    digits = re.sub(r"[^\d]", "", str(value))
    if not digits:
        return None
    try:
        numeric = int(digits)
    except ValueError:
        return None
    return numeric if numeric > 0 else None


def price_from_option_data(option_data):
    if not option_data:
        return None
    candidates = []
    for data in option_data.values():
        prices = data.get('하위옵션가격') if isinstance(data, dict) else None
        if not prices:
            continue
        for raw in prices:
            normalized = normalize_price_value(raw)
            if normalized:
                candidates.append(normalized)
    if not candidates:
        return None
    return min(candidates)


_PRELOADED_PRICE_SCRIPT = """
() => {
    const state = window.__PRELOADED_STATE__;
    if (!state || !state.productSimpleView || !state.productSimpleView.product) {
        return null;
    }
    const product = state.productSimpleView.product;
    const wrap = (value) => {
        const type = typeof value;
        if (type === "number" || type === "string") {
            return value;
        }
        return null;
    };
    return {
        salePrice: wrap(product.salePrice),
        discountedSalePrice: wrap(product.discountedSalePrice),
        price: wrap(product.price)
    };
}
"""


def price_from_preloaded_state(page):
    try:
        price_info = page.evaluate(_PRELOADED_PRICE_SCRIPT)
    except Exception as exc:
        print(f"Failed to read PRELOADED_STATE price: {exc}")
        return None
    if not price_info:
        return None
    for key in ("salePrice", "discountedSalePrice", "price"):
        normalized = normalize_price_value(price_info.get(key))
        if normalized:
            return normalized
    return None


def update_query_params(url, **params):
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({key: value for key, value in params.items() if value is not None})
    new_query = urlencode(query, doseq=True)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))


class Tee(object):
    def __init__(self, *files):
        self.files = files

    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()

    def flush(self):
        for f in self.files:
            f.flush()


f = LOG_FILE.open('w', encoding='utf-8')
original = sys.stdout
sys.stdout = Tee(sys.stdout, f)


base_url = "https://smartstore.naver.com"


def product_list_crawl(context, df, read_excel_path, seen_urls):
    page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(page)

    raw_url = 'https://smartstore.naver.com/joypapa_/category/ALL?st=RECENT&dt=BIG_IMAGE&size=20'
    original_url = update_query_params(raw_url, page=None)
    page.goto(original_url)
    page.wait_for_load_state("load")
    page.wait_for_load_state("networkidle")

    global_start_page = 61
    global_last_page = 61
    # 디버그: 특정 페이지만 요청된 경우 범위를 해당 값으로 축소
    if CRAWL_ONLY_PAGES:
        try:
            global_start_page = min(CRAWL_ONLY_PAGES)
            global_last_page = max(CRAWL_ONLY_PAGES)
        except Exception:
            pass
    shopname = raw_url.split('/')[3]
    shopnumber = raw_url.split('/')[5].split('?')[0]

    home_dir = Path.home()
    output_folder = home_dir / 'Desktop' / 'excel_output'
    output_folder.mkdir(parents=True, exist_ok=True)

    pagination_button_labels = {
        "next": ["다음", "다음 페이지", "다음페이지", ">"],
        "prev": ["이전", "이전 페이지", "이전페이지", "<"]
    }
    reached_total_limit = False

    # 검증 모드: 특정 페이지의 첫 상품을 열어 기대 URL/이름 확인
    verify_target_page = None
    verify_expected_url = (os.getenv("VERIFY_FIRST_PRODUCT_URL") or "").strip() or None
    verify_expected_name = (os.getenv("VERIFY_FIRST_PRODUCT_NAME") or "").strip() or None
    try:
        verify_target_page = int(os.getenv("VERIFY_TARGET_PAGE", "") or 0) or None
    except Exception:
        verify_target_page = None

    def scroll_to_pagination():
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        except PlaywrightTimeoutError:
            pass
        time.sleep(0.8)
        # 하단 스크롤 후 간단 스크린샷
        debug_shot(page, "scrolled_bottom")

    def wait_pagination_ready(timeout_ms=8000):
        # 페이지네이션 컨테이너 또는 숫자/버튼이 나타날 때까지 대기
        selectors = [
            "div[data-shp-area='list.pgn'][role='menubar']",
            "div[data-shp-contents-type='pgn'][role='menubar']",
            "div[data-shp-area-id='pgn'][role='menubar']",
        ]
        end = time.time() + (timeout_ms / 1000.0)
        while time.time() < end:
            container = find_pagination_container()
            if container:
                try:
                    if container.query_selector("a[role='menuitem'],a[role='button']"):
                        return True
                except Exception:
                    pass
            time.sleep(0.3)
        return False

    def get_first_list_href():
        try:
            el = page.query_selector("a[href*='/products/']")
            if el:
                href = el.get_attribute("href")
                return href
        except Exception:
            pass
        return None

    def find_pagination_container():
        candidates = [
            "div[data-shp-area='list.pgn'][role='menubar']",
            "div[data-shp-contents-type='pgn'][role='menubar']",
            "div[data-shp-area-id='pgn'][role='menubar']",
            # 폴백들
            "nav[aria-label*='페이지']",
            "nav[aria-label*='pagination']",
            "nav[role='navigation']",
            "div[class*='Pagination']",
            "div[class*='paginate']",
            "div[class*='paging']",
        ]
        for sel in candidates:
            try:
                elem = page.query_selector(sel)
            except Exception:
                elem = None
            if elem:
                return elem
        return None

    def get_pgn_from_container():
        """컨테이너의 data-shp-filter_con 속성에서 pgn 값을 파싱(네이버 구조 특화)."""
        try:
            container = find_pagination_container()
            if not container:
                return None
            raw = container.get_attribute("data-shp-filter_con")
            if not raw:
                return None
            # HTML 인코딩된 문자열 처리
            raw = raw.replace('&quot;', '"')
            data = json.loads(raw)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict) and item.get('key') in {'pgn', 'page', 'pageNum'}:
                        val = item.get('value')
                        try:
                            return int(re.findall(r"\d+", str(val))[0])
                        except Exception:
                            pass
        except Exception:
            return None
        return None

    def find_page_link_in_container(target_page):
        container = find_pagination_container()
        if container is None:
            return None
        # 접근성 역할 기반으로 우선 탐색
        try:
            cand = container.get_by_role("link", name=str(target_page))
            if cand and cand.count() > 0:
                return cand.first
        except Exception:
            pass
        try:
            cand = container.get_by_role("button", name=str(target_page))
            if cand and cand.count() > 0:
                return cand.first
        except Exception:
            pass
        # 숫자 텍스트 필터 탐색
        for sel in ["a", "button", "span", "li"]:
            for node in container.query_selector_all(sel):
                try:
                    t = (node.inner_text() or "").strip()
                except Exception:
                    continue
                if t.isdigit() and int(t) == target_page:
                    return node
        return None

    def ensure_group_has_page(target_page, max_group_hops=20):
        """타겟 숫자 링크가 현재 보이는 그룹에 나타나도록 '다음' 그룹 이동을 반복."""
        hops = 0
        while hops < max_group_hops:
            scroll_to_pagination()
            link = find_page_link_in_container(target_page)
            if link:
                return True
            # 다음 그룹 이동 시도
            before_sig = get_first_list_href()
            print(f"타겟 {target_page}가 보이지 않아 '다음' 그룹 이동 시도")
            if not click_pagination_control("next"):
                break
            # 리스트 변화 대기 (첫 상품 href 변경 기준)
            for _ in range(10):
                time.sleep(0.5)
                after_sig = get_first_list_href()
                if before_sig and after_sig and before_sig != after_sig:
                    break
            hops += 1
        return False

    def get_current_page_number():
        # 0) 컨테이너의 pgn 값 파싱 시도
        pgn_val = get_pgn_from_container()
        if isinstance(pgn_val, int):
            return pgn_val
        selectors = [
            'a[aria-current="true"]',
            'button[aria-current="true"]',
            '[aria-current="page"]'
        ]
        # 컨테이너에서 role=menuitem + aria-current 우선 확인
        try:
            container = find_pagination_container()
            if container:
                node = container.query_selector("a[role='menuitem'][aria-current='true']")
                if node:
                    txt = (node.inner_text() or '').strip()
                    m = re.search(r'\d+', txt)
                    if m:
                        return int(m.group(0))
        except Exception:
            pass
        for selector in selectors:
            locator = page.locator(selector)
            try:
                locator.first.wait_for(state="attached", timeout=5000)
            except PlaywrightTimeoutError:
                continue

            try:
                text = locator.first.inner_text().strip()
            except PlaywrightTimeoutError:
                continue

            match = re.search(r'\d+', text)
            if match:
                return int(match.group())
        # 쿼리스트링에서 page 파라미터 추출 시도
        try:
            from urllib.parse import urlsplit, parse_qsl
            parts = urlsplit(page.url)
            query = dict(parse_qsl(parts.query, keep_blank_values=True))
            for key in ["page", "pageIndex", "pagingIndex", "pageNum", "p"]:
                if key in query:
                    try:
                        return int(re.findall(r"\d+", query[key])[0])
                    except Exception:
                        pass
        except Exception:
            pass
        print(f"현재 페이지 번호 탐색 실패 - URL: {page.url}")
        try:
            print("aria-current 후보:", page.locator('[aria-current]').all_inner_texts())
        except Exception:
            pass
        return None

    def find_page_link(target_page):
        # 1) 기존 네비게이션 링크(a[role=menuitem])에서 검색
        for link in page.query_selector_all('a[role="menuitem"]'):
            try:
                text = link.inner_text().strip()
            except PlaywrightTimeoutError:
                continue
            match = re.search(r'^\d+$', text)
            if match and int(text) == target_page:
                return link

        # 2) 접근성 역할 기반 탐색
        try:
            candidate = page.get_by_role("link", name=str(target_page))
            if candidate and candidate.count() > 0:
                return candidate.first
        except Exception:
            pass
        try:
            candidate = page.get_by_role("button", name=str(target_page))
            if candidate and candidate.count() > 0:
                return candidate.first
        except Exception:
            pass

        # 3) 포괄적 탐색(a, button)에서 텍스트가 숫자만이고 타겟과 일치하는 요소 선택
        for sel in ["a", "button"]:
            for link in page.query_selector_all(sel):
                try:
                    text = link.inner_text().strip()
                except Exception:
                    continue
                if not text:
                    continue
                if re.fullmatch(r"\d+", text) and int(text) == target_page:
                    return link

        return None

    def verify_first_product_on_page():
        try:
            # 첫 상품 링크 탐색
            page.wait_for_selector("a[href*='/products/']", timeout=10000)
            first_link = page.query_selector("a[href*='/products/']")
            if not first_link:
                print("검증 실패: 첫 상품 링크를 찾지 못했습니다.")
                return False
            href = first_link.get_attribute("href") or ""
            first_link.click()
            page.wait_for_load_state("load")
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError:
                pass
            final_url = page.url
            print(f"검증용 이동 URL: {final_url}")
            ok_url = True
            if verify_expected_url:
                ok_url = (verify_expected_url in final_url)
            ok_name = True
            if verify_expected_name:
                try:
                    content_text = page.inner_text("body")
                except Exception:
                    content_text = ""
                ok_name = (verify_expected_name in content_text)
            if ok_url and ok_name:
                print("검증 성공: 기대 URL/이름과 일치합니다.")
                return True
            else:
                print(f"검증 결과: URL일치={ok_url}, 이름일치={ok_name}")
                return False
        except Exception as exc:
            print(f"검증 중 예외: {exc}")
            return False

    def click_pagination_control(direction):
        labels = pagination_button_labels[direction]
        for label in labels:
            try:
                page.get_by_role("button", name=label).click(timeout=1500)
                page.wait_for_load_state("networkidle")
                time.sleep(1)
                return True
            except PlaywrightTimeoutError:
                continue
            except Exception:
                continue

        # 컨테이너 내부에서 '다음/이전' 우선 탐색
        container = find_pagination_container()
        if container:
            try:
                # role=button + aria-hidden=false 후보들 검사
                for node in container.query_selector_all("a[role='button'],button[role='button']"):
                    try:
                        hidden = node.get_attribute('aria-hidden')
                    except Exception:
                        hidden = None
                    if hidden == 'true':
                        continue
                    try:
                        t = (node.inner_text() or '').strip()
                    except Exception:
                        t = ''
                    if direction == 'next' and ("다음" in t or t in {"›", ">", "»"}):
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                    if direction == 'prev' and ("이전" in t or t in {"‹", "<", "«"}):
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                # 텍스트 기반 후보
                text_sel = (
                    "a:has-text('다음'),button:has-text('다음')" if direction == 'next' else "a:has-text('이전'),button:has-text('이전')"
                )
                cand = container.query_selector(text_sel)
                if cand:
                    try:
                        cand.scroll_into_view_if_needed()
                    except Exception:
                        pass
                    cand.click()
                    page.wait_for_load_state("networkidle")
                    time.sleep(1)
                    return True
                # 구조 기반 폴백: role=button 앵커 배열의 양 끝을 사용
                rb = container.query_selector_all("a[role='button'],button[role='button']")
                if rb:
                    try:
                        node = rb[-1] if direction == 'next' else rb[0]
                        try:
                            node.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                    except Exception:
                        pass
                # data-shp-contents-id 보유 요소 우선 클릭(네이버 특화)
                rb2 = container.query_selector_all("a[role='button'][data-shp-contents-id],button[role='button'][data-shp-contents-id]")
                if rb2:
                    try:
                        node = rb2[-1] if direction == 'next' else rb2[0]
                        try:
                            node.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        node.click()
                        page.wait_for_load_state("networkidle")
                        time.sleep(1)
                        return True
                    except Exception:
                        pass
            except Exception:
                pass

        selector = (
            'a[role="button"][aria-hidden="false"]:last-child'
            if direction == "next"
            else 'a[role="button"][aria-hidden="false"]:first-child'
        )
        button = page.query_selector(selector)
        if button:
            button.click()
            page.wait_for_load_state("networkidle")
            time.sleep(1)
            return True

        return False

    def go_to_page_number(target_page):
        attempt = 0
        max_attempts = 30
        # next_only 전략: 숫자 링크 사용 없이 '다음'만 반복 클릭
        if PAGINATION_STRATEGY == 'next_only':
            scroll_to_pagination()
            wait_pagination_ready(8000)
            cur = get_current_page_number() or 1
            target = int(target_page)
            print(f"next_only: 현재 {cur} → 목표 {target}")
            # 안전 범위 내에서 목표까지 전진
            for _ in range(min(200, max(0, target - cur) + 20)):
                if cur >= target:
                    break
                before_sig = get_first_list_href()
                if not click_pagination_control('next'):
                    print("next 버튼 클릭 실패")
                    return False
                for __ in range(12):
                    time.sleep(0.5)
                    after_sig = get_first_list_href()
                    if before_sig and after_sig and before_sig != after_sig:
                        break
                # aria-current가 없을 수 있으므로 보수적으로 증가
                new_cur = get_current_page_number()
                cur = new_cur if new_cur is not None else (cur + 1)
            if cur == target:
                print(f"next_only: 페이지 {target} 도달")
                return True
            print("next_only: 페이지 번호 미판별, 리스트 변화 기준 성공 처리")
            return True
        while attempt < max_attempts:
            attempt += 1
            scroll_to_pagination()
            current_page_num = get_current_page_number()
            debug_shot(page, f"attempt{attempt}_target{target_page}_after_detect")

            if current_page_num == target_page:
                print(f"페이지 {target_page}에 이미 위치해 있습니다.")
                return True

            # 먼저 현재 보이는 그룹 내에서 타겟 숫자 링크를 찾음
            page_link = find_page_link_in_container(target_page) or find_page_link(target_page)
            if page_link:
                try:
                    page_link.scroll_into_view_if_needed()
                except PlaywrightTimeoutError:
                    pass
                time.sleep(0.3)
                debug_shot(page, f"attempt{attempt}_target{target_page}_before_link_click")
                before_sig = get_first_list_href()
                page_link.click()
                page.wait_for_load_state("networkidle")
                time.sleep(1)
                debug_shot(page, f"attempt{attempt}_target{target_page}_after_link_click")
                if get_current_page_number() == target_page:
                    print(f"페이지 {target_page}로 이동 완료, 현재 URL: {page.url}")
                    return True
                # 페이지 번호 판단이 불가한 경우, 리스트 시그니처 변경으로 이동 검증
                after_sig = get_first_list_href()
                if before_sig and after_sig and before_sig != after_sig:
                    print(f"리스트 변경 감지로 페이지 {target_page} 이동 성공으로 간주")
                    return True
                continue

            if current_page_num is None:
                print("현재 페이지 번호를 확인할 수 없어 다시 시도합니다.")
                # URL 파라미터 기반 점프를 우선 1회 시도
                if PAGE_JUMP_BY_QUERY and attempt in {1, 5, 10, 20}:
                    try:
                        for key in ["page", "pageIndex", "pagingIndex", "pageNum", "p"]:
                            jump_url = update_query_params(original_url, **{key: target_page})
                            print(f"URL 점프 시도(번호 미탐지): {jump_url}")
                            page.goto(jump_url)
                            page.wait_for_load_state("networkidle")
                            time.sleep(1)
                            num = get_current_page_number()
                            if num == target_page:
                                print(f"URL 점프로 페이지 {target_page} 이동 확인")
                                return True
                    except Exception as exc:
                        print(f"URL 점프 실패: {exc}")
                # 숫자 링크가 보이는 그룹이 아닐 수 있으니 그룹 이동 시도
                if ensure_group_has_page(target_page):
                    continue
                page.wait_for_timeout(1000)
                continue

            direction = "next" if target_page > current_page_num else "prev"
            print(f"페이지 {target_page} 이동을 위해 {direction} 버튼 클릭 시도 (현재 {current_page_num}).")
            before_sig = get_first_list_href()
            if not click_pagination_control(direction):
                print(f"{direction} 버튼을 찾을 수 없습니다.")
                # 버튼 탐색 실패 시 URL 파라미터 기반 점프 시도
                if PAGE_JUMP_BY_QUERY:
                    try:
                        for key in ["page", "pageIndex", "pagingIndex", "pageNum", "p"]:
                            jump_url = update_query_params(original_url, **{key: target_page})
                            print(f"URL 점프 시도: {jump_url}")
                            debug_shot(page, f"attempt{attempt}_target{target_page}_before_url_jump")
                            page.goto(jump_url)
                            page.wait_for_load_state("networkidle")
                            time.sleep(1)
                            debug_shot(page, f"attempt{attempt}_target{target_page}_after_url_jump")
                            num = get_current_page_number()
                            if num == target_page:
                                print(f"URL 점프로 페이지 {target_page} 이동 확인")
                                return True
                    except Exception as exc:
                        print(f"URL 점프 실패: {exc}")
                return False
            # 리스트 변경으로 이동 검증
            for _ in range(10):
                time.sleep(0.5)
                after_sig = get_first_list_href()
                if before_sig and after_sig and before_sig != after_sig:
                    break

        print(f"페이지 {target_page} 이동 시도가 {max_attempts}회 초과로 실패했습니다.")
        return False

    # 지정된 페이지만 크롤링하도록 제한(있을 경우)
    only_pages_set = set(CRAWL_ONLY_PAGES) if CRAWL_ONLY_PAGES else None

    for start_page in range(global_start_page, global_last_page + 1, 10):
        last_page = min(start_page + 9, global_last_page)

        # 이 그룹(10페이지 묶음)에 처리할 페이지가 없으면 건너뜀
        if only_pages_set is not None:
            group_pages = set(range(start_page, last_page + 1))
            group_target_pages = sorted(group_pages & only_pages_set)
            if not group_target_pages:
                print(f"Skip page group {start_page}-{last_page} (no target pages in CRAWL_ONLY_PAGES)")
                continue
        else:
            group_target_pages = list(range(start_page, last_page + 1))
        write_excel_path = output_folder / f'dolce_{shopname}_{shopnumber}_{start_page}_{last_page}.xlsx'
        shutil.copy(read_excel_path, write_excel_path)

        # 그룹 내 최초 타겟 페이지로 이동
        first_target = group_target_pages[0]
        if not go_to_page_number(first_target):
            print(f"페이지 {start_page} 이동에 실패했습니다. 다음 그룹으로 넘어갑니다.")
            continue

        for page_number in group_target_pages:
            if not go_to_page_number(page_number):
                print(f"페이지 {page_number} 이동에 실패하여 건너뜁니다.")
                continue
            # 검증 모드: 대상 페이지에서 첫 상품 열어 확인 후 종료
            if verify_target_page and page_number == verify_target_page:
                ok = verify_first_product_on_page()
                print(f"VERIFY_RESULT: page={page_number}, ok={ok}")
                return
            df, _ = crawl_page(page, df, seen_urls)
            print(f"Completed page {page_number}")
            if MAX_PRODUCTS_TOTAL and len(df) >= MAX_PRODUCTS_TOTAL:
                df = df.iloc[:MAX_PRODUCTS_TOTAL]
                reached_total_limit = True
                print(f"Reached MAX_PRODUCTS_TOTAL={MAX_PRODUCTS_TOTAL}, stopping after page {page_number}.")
                break

        write_to_excel(df, write_excel_path, seen_urls)
        write_to_excel2(
            df,
            output_folder / f'dolce_{shopname}_{shopnumber}_{start_page}_{last_page}_second.xlsx'
        )
        print(f"Processed pages {start_page} to {last_page}")
        if reached_total_limit:
            print("MAX_PRODUCTS_TOTAL reached; ending crawl.")
            break

    page.close()


def ensure_product_detail_visible(page):
    """Ensure the SmartStore 상세정보 영역 is expanded so selectors become available."""
    toggle_selectors = [
        "button[data-resize-on-click='true']",
        "button:has-text('상세정보 펼치기')",
        "button:has-text('상세정보 더보기')",
    ]
    for selector in toggle_selectors:
        try:
            toggle = page.query_selector(selector)
        except Exception:
            toggle = None
        if not toggle:
            continue
        try:
            aria_expanded = (toggle.get_attribute("aria-expanded") or "").lower()
        except Exception:
            aria_expanded = ""
        try:
            label = (toggle.inner_text() or "").strip()
        except Exception:
            label = ""
        need_expand = (
            aria_expanded == "false"
            or ("펼치기" in label and "접기" not in label)
            or ("더보기" in label and "접기" not in label)
        )
        if need_expand:
            try:
                toggle.scroll_into_view_if_needed()
            except Exception:
                pass
            toggle.click()
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(500)
        break
    for _ in range(4):
        try:
            section = page.query_selector("#INTRODUCE")
        except Exception:
            section = None
        if section:
            try:
                section.scroll_into_view_if_needed()
            except Exception:
                try:
                    page.evaluate("document.getElementById('INTRODUCE')?.scrollIntoView({behavior: 'instant', block: 'start'})")
                except Exception:
                    pass
            try:
                section.wait_for_element_state("visible", timeout=2000)
            except Exception:
                pass
            break
        try:
            page.mouse.wheel(0, 1200)
        except Exception:
            try:
                page.evaluate("window.scrollBy(0, document.body.scrollHeight / 3)")
            except Exception:
                pass
        page.wait_for_timeout(500)
    try:
        page.wait_for_selector("#INTRODUCE", timeout=5000)
    except Exception:
        pass


def find_content_element(page, product_code):
    time.sleep(1)
    ensure_product_detail_visible(page)

    content_selectors = [
        '#INTRODUCE > div > div.LXGzUhHJC2.EtTm8LLHdw.Uea3oKmnaJ > div > div > div > div > div > div > div',
        '#INTRODUCE > div > div.LXGzUhHJC2.EtTm8LLHdw > div > div > div > div > div > div > div',
        '#INTRODUCE .detail_viewer',
        '#INTRODUCE [data-component-id]',
        '#INTRODUCE .se-main-container',
        '#INTRODUCE',
        '[data-name="INTRODUCE"][role="tabpanel"]',
        'xpath=//*[@id="INTRODUCE"]//div[contains(@data-component-id,"INTRODUCE")]//div[contains(@class,"se_component")]//div[last()]',
        'xpath=//*[@id="INTRODUCE"]//div[contains(@class,"se-main-container")]',
        'xpath=//*[@id="INTRODUCE"]/div/div[4]',
    ]

    for selector in content_selectors:
        print(f"[CONTENT][{product_code}] Trying selector: {selector}")
        element = page.query_selector(selector)
        if element is not None:
            print(f"Using content selector '{selector}' for {product_code}")
            return selector

    print(f"[CONTENT][{product_code}] 상품 상세 컨텐츠 영역을 찾지 못했습니다. 스냅샷 저장 후 None 반환")
    save_debug_snapshot(page, f"content_{product_code}")
    return None


def crawl_page(page, df, seen_urls):
    time.sleep(1)
    page.wait_for_load_state("networkidle")
    # 상품 링크 등장 대기 (동적 로딩 대비)
    try:
        page.wait_for_selector("a[href*='/products/']", timeout=10000)
    except Exception:
        pass
    products = find_elements(
        page,
        [
            "[data-testid='PRODUCT_CARD']",
            "li:has(a[href*='/products/'])",
            "div:has(a[href*='/products/'])",
            "li[class*='flu7YgFW2k']",
        ],
    )
    if not products:
        print("상품 리스트 셀렉터가 모두 실패했습니다. HTML 스냅샷을 저장합니다.")
        save_debug_snapshot(page, "product_list")
    duplicate_detected = False

    for i, product in enumerate(products):
        product_data = get_product_data(page, product, i, len(products))
        if product_data is None:
            print(f"Skipping product at index {i} as get_product_data returned None.")
            continue

        product_url = product_data['Product_URL'][0]

        if product_url == "N/A" or not product_url:
            print("상품 URL 추출 실패로 항목을 건너뜁니다.")
            continue

        if product_url in seen_urls:
            print('Duplicate product detected: ', product_url)
            duplicate_detected = True
            break
        else:
            seen_urls.add(product_url)

        df = pd.concat([df, product_data], ignore_index=True)

        if MAX_PRODUCTS_PER_PAGE and (i + 1) >= MAX_PRODUCTS_PER_PAGE:
            print(f"Reached MAX_PRODUCTS_PER_PAGE={MAX_PRODUCTS_PER_PAGE}, stop crawling this page.")
            break

    return df, duplicate_detected


def extract_product_details(product):
    title_element = first_available(
        product,
        [
            "strong[aria-hidden='false']",
            "[data-testid='PRODUCT_CARD_TITLE']",
            "a[href*='/products/'] strong",
            "span[class*='ProductCard__Title']",
            "strong._26YxgX-Nu5",
        ],
    )
    if title_element:
        title = title_element.inner_text().strip()
    else:
        title = product.inner_text().splitlines()[0].strip() if product.inner_text() else "N/A"

    price_element = first_available(
        product,
        [
            "[data-testid='PRODUCT_CARD_PRICE']",
            "span:has-text('원')",
            "strong span:has-text('원')",
            "span._2DywKu0J_8",
        ],
    )
    if price_element:
        price = extract_price_from_text(price_element.inner_text())
    else:
        price = extract_price_from_text(product.inner_text())

    url_element = first_available(
        product,
        [
            "a[href*='/products/'][role='link']",
            "a[href*='/products/']",
            "a._2id8yXpK_k",
        ],
    )

    if url_element:
        raw_url = url_element.get_attribute("href")
        if raw_url and raw_url.startswith("/"):
            product_url = base_url + raw_url
        else:
            product_url = raw_url
    else:
        product_url = None

    product_code = product_url.split('/')[-1] if product_url else "N/A"
    return title, price, product_url or "N/A", product_code


def original_shipping_fee(page):
    print(f"Current page URL: {page.url}")

    shipping_selectors = [
        "xpath=//*[contains(@class,'delivery') and contains(text(),'원')]",
        "xpath=//span[contains(text(),'배송비')]/following-sibling::*[1]",
        "xpath=//*[contains(text(),'배송비') and contains(text(),'원')]",
        "xpath=//*[contains(text(),'반품배송비') and contains(text(),'원')]",
    ]

    for selector in shipping_selectors:
        element = page.query_selector(selector)
        if not element:
            continue
        element_text = element.inner_text().strip()
        if "무료배송" in element_text:
            print("배송비: 무료배송")
            return "0"
        digits = re.findall(r"[\d,]+", element_text)
        if digits:
            value = digits[0].replace(",", "")
            print(f"배송비: {value}")
            return value

    body_text = ""
    try:
        body_text = page.inner_text("body")
    except Exception:
        pass

    if "무료배송" in body_text:
        print("배송비: 무료배송(본문 탐지)")
        return "0"

    for pattern in [r"배송비\s*[:：]?\s*([\d,]+)\s*원", r"반품배송비\s*[:：]?\s*([\d,]+)\s*원"]:
        match = re.search(pattern, body_text)
        if match:
            value = match.group(1).replace(",", "")
            print(f"배송비(본문 탐지): {value}")
            return value

    print("Shipping fee element not found, 저장 후 N/A 반환")
    save_debug_snapshot(page, "shipping_fee")
    return "N/A"


def option_crawl(page):
    option_data = {}

    option_triggers = page.query_selector_all('[data-shp-area$="optselect"]')
    if not option_triggers:
        option_triggers = page.query_selector_all(
            'a[role="button"][aria-haspopup="listbox"], button[aria-haspopup="listbox"]'
        )
        if option_triggers:
            print("Fallback option selector 사용 (listbox 버튼 기반)")

    option_index = 0
    while True:
        current_triggers = page.query_selector_all('[data-shp-area$="optselect"]')
        if not current_triggers:
            current_triggers = option_triggers

        if option_index >= len(current_triggers):
            break

        trigger = current_triggers[option_index]
        data_area = (trigger.get_attribute("data-shp-area") or "")
        if data_area and "optselect" not in data_area:
            option_index += 1
            continue

        option_index += 1
        category = trigger.get_attribute("aria-label") or trigger.inner_text().strip()
        if not category or category in {"선택", ""}:
            category = f"옵션{option_index}"

        try:
            trigger.click()
        except PlaywrightTimeoutError:
            print(f"{category} 클릭 실패")
            continue

        try:
            dropdown = page.wait_for_selector("ul[role=\"listbox\"]", timeout=15000)
        except PlaywrightTimeoutError:
            print(f"{category} 옵션 리스트 로드 실패")
            continue

        items = dropdown.query_selector_all("[role='option'], a, li")
        current_options = []
        current_prices = []

        for item in items:
            option_text = item.inner_text().strip()
            if not option_text:
                continue
            price_match = re.search(r'\(([+\-]?[\d,]+)원\)', option_text)
            if price_match:
                price_value = int(price_match.group(1).replace(',', ''))
                name = re.sub(r'\(([+\-]?[\d,]+)원\)', '', option_text).strip()
            else:
                price_value = 0
                name = option_text

            current_options.append(name)
            current_prices.append(price_value)

        if not current_options:
            print(f"{category} 옵션 정보를 찾지 못했습니다.")
            continue

        option_data[category] = {
            '하위옵션제목': current_options,
            '하위옵션가격': current_prices,
        }

        selectable = []
        for item in items:
            try:
                disabled = item.evaluate("node => node.getAttribute('aria-disabled') === 'true'")
            except Exception:
                disabled = False
            if not disabled:
                selectable.append(item)

        if selectable:
            random.choice(selectable).click()
            time.sleep(0.5)

    return option_data


def image_crawl(page):
    main_candidates = find_elements(
        page,
        [
            "img[alt='대표이미지']",
            "img[alt*='대표'][src*='shop-phinf']",
            "div[id='content'] img[src*='shop-phinf']",
        ],
    )
    thumbnail_elements = find_elements(
        page,
        [
            "img[alt^='추가이미지']",
            "button[aria-label^='썸네일'] img",
            "ul[class*='thumbnail'] img",
        ],
    )

    if not thumbnail_elements:
        thumbnail_elements = page.query_selector_all("img[src*='shop-phinf']")

    image_elements = []
    if main_candidates:
        image_elements.extend(main_candidates)
    if thumbnail_elements:
        image_elements.extend(thumbnail_elements)

    if not image_elements:
        try:
            fallback = page.wait_for_selector(
                'xpath=//*[@id="content"]//img[contains(@src,"shop-phinf")]',
                timeout=5000,
            )
            if fallback:
                image_elements = [fallback]
        except Exception:
            pass

    if not image_elements:
        print("No images found on the page.")
        return [], []

    thumbnail_urls = []
    for element in image_elements:
        src = element.get_attribute("src")
        if not src:
            continue
        thumbnail_urls.append(src.split("?")[0])

    seen = set()
    unique_urls = []
    for url in thumbnail_urls:
        if url in seen:
            continue
        seen.add(url)
        unique_urls.append(url)

    if not unique_urls:
        print("Image URLs could not be extracted.")
        return [], []

    def url_prefix(u):
        filename = u.split("/")[-1]
        digits = "".join(ch for ch in filename if ch.isdigit())
        return digits[:3] if digits else filename[:3]

    prefixes = [url_prefix(url) for url in unique_urls]
    counts = Counter(prefixes)
    if not counts:
        return unique_urls, []

    most_common = counts.most_common(1)[0][0]
    common_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix == most_common]
    different_urls = [url for url, prefix in zip(unique_urls, prefixes) if prefix != most_common]

    return common_urls, different_urls


BRANDING_IMAGE_URLS = {
    "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/top.png",
    "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/bottom.png",
    "https://coudae.s3.ap-northeast-2.amazonaws.com/A00412936/cloud/7290.png",
}

_GRAY_LINE_REPLACEMENTS = {
    "https://rapid-up.s3.ap-northeast-2.amazonaws.com/dev/gray-line.png":
        "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/gray-line.png"
}

_BLOCKED_IMAGE_PREFIXES = (
    "https://rapid-up.s3.ap-northeast-2.amazonaws.com",
    "https://cdn.heyseller.kr",
    "https://ai.esmplus.com/",
)


def log_content_debug(product_code, message):
    print(f"[CONTENT][{product_code}] {message}")


def wrap_html_document(snippet):
    if not snippet:
        return ""
    return (
        "<!DOCTYPE html>"
        "<html lang=\"ko\">"
        "<head>"
        "<meta charset=\"utf-8\">"
        "<style>body{margin:0;padding:0;background:#fff;}</style>"
        "</head>"
        "<body>"
        f"{snippet}"
        "</body>"
        "</html>"
    )


def dump_content_html(product_code, html_text, label):
    if not DUMP_CONTENT_HTML:
        return
    if not html_text:
        return
    try:
        DUMP_CONTENT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = DUMP_CONTENT_DIR / f"{product_code}_{label}_{timestamp}.html"
        filename.write_text(html_text, encoding="utf-8")
        print(f"[CONTENT][{product_code}] Saved content output ({label}): {filename}")
    except Exception as exc:
        print(f"[CONTENT][{product_code}] Failed to save content output: {exc}")


def cleanup_dom_structure(soup, product_code):
    removed = 0

    for tag_name in ("style", "script", "svg", "canvas"):
        for node in list(soup.find_all(tag_name)):
            node.decompose()
            removed += 1

    selectors_to_remove = [
        ".pzp-ui-dimmed",
        ".pzp-upnext-endscreen",
        ".pzp-ui-playlist",
        ".pzp-double-tap-overlay",
        ".pzp-ad-break-indicator",
        ".pzp-pc__poster",
        ".pzp-ui-circle-process",
        ".pzp-ui-dimmed",
    ]
    for selector in selectors_to_remove:
        for node in list(soup.select(selector)):
            node.decompose()
            removed += 1

    selectors_to_unwrap = [
        ".se-main-container",
        ".editor_wrap",
        ".se-viewer",
        ".uOXg8u0yzs",
        ".LXGzUhHJC2",
        ".EtTm8LLHdw",
        ".Uea3oKmnaJ",
        ".se-component",
        ".se-component-content",
        ".se-section",
        ".se-module",
        ".se-section-video",
        ".se-module-video",
        "[aria-label*='비디오']",
        "[aria-label*='동영상']",
        "[id^='wpc-']",
        ".pzp-pc__video",
    ]
    for selector in selectors_to_unwrap:
        for node in list(soup.select(selector)):
            node.unwrap()

    if removed:
        log_content_debug(product_code, f"Removed {removed} extra DOM nodes during cleanup.")


def strip_blob_media(soup, product_code):
    removed = 0
    simplified_videos = 0

    for component in list(soup.select(".se-component.se-video")):
        playable_src = None
        video_tags = component.find_all("video")
        for video_tag in video_tags:
            src_candidates = [video_tag.get("src"), video_tag.get("data-src")]
            for candidate in src_candidates:
                if candidate and not candidate.startswith("blob:"):
                    playable_src = candidate
                    break
            if playable_src:
                break

        if playable_src:
            simple_video = soup.new_tag("video")
            simple_video["src"] = playable_src
            simple_video["controls"] = "controls"
            simple_video["autoplay"] = "autoplay"
            simple_video["muted"] = "muted"
            simple_video["loop"] = "loop"
            simple_video["style"] = "display:block;margin:0 auto 20px auto;max-width:100%;"
            component.replace_with(simple_video)
            simplified_videos += 1
        else:
            component.decompose()
            removed += 1

    blob_tags = []
    for tag in soup.find_all(["video", "source", "iframe", "canvas"]):
        attrs = [
            tag.get("src", ""),
            tag.get("data-src", ""),
            tag.get("poster", ""),
        ]
        if any(attr and "blob:" in attr for attr in attrs):
            blob_tags.append(tag)

    for tag in blob_tags:
        parent = tag.find_parent(class_="se-component se-video") or tag
        parent.decompose()
        removed += 1

    extra_selectors = [
        ".prismplayer-area",
        "[class*='pzp-']",
        "[class*='pzp_']",
        "[class*='pzp ']",
        "[class~='pzp']",
        "[class*='webplayer']",
        "[class*='player-area']",
        "[class*='pzp-pc']",
        "[class*='pzp-ui']",
        "[class*='pzp-upnext']",
    ]
    for selector in extra_selectors:
        for node in soup.select(selector):
            node.decompose()
            removed += 1

    video_text_keywords = (
        "광고 후 계속됩니다",
        "다음 동영상",
        "subject",
        "author",
        "재생 속도",
        "해상도",
        "자막",
        "옵션",
        "도움말",
        "죄송합니다. 문제가 발생했습니다",
        "고화질 재생이 가능한 영상입니다",
        "더 알아보기",
        "00:00",
        "0:00",
    )
    removed_text = 0
    for text_node in list(soup.find_all(string=True)):
        stripped = text_node.strip()
        if not stripped:
            continue
        if any(keyword in stripped for keyword in video_text_keywords):
            container = getattr(text_node, "parent", None)
            if container is None:
                continue
            if container.name in {"html", "body"}:
                continue
            parent_component = container.find_parent(class_="se-component se-video")
            target = parent_component or container
            try:
                target.decompose()
                removed_text += 1
            except Exception:
                continue
    removed += removed_text

    if removed or simplified_videos:
        log_content_debug(
            product_code,
            f"Removed {removed} blob media blocks, simplified {simplified_videos} playable videos.",
        )


def content_crawl(page, product_code, element_selector):
    time.sleep(1)
    page.wait_for_load_state("load")

    if not element_selector:
        log_content_debug(product_code, "No element selector available.")
        return None

    element = page.query_selector(element_selector)
    if element is None:
        log_content_debug(product_code, f"Selector '{element_selector}' resolved to None.")
        save_debug_snapshot(page, f"content_missing_{product_code}")
        return None

    raw_content = element.inner_html()
    if not (raw_content or "").strip():
        log_content_debug(product_code, "Element inner_html is empty; capturing page snapshot.")
        save_debug_snapshot(page, f"content_empty_{product_code}")
        return None
    soup = BeautifulSoup(raw_content, 'html.parser')

    text_snapshot = soup.get_text(strip=True)
    normalized_text = text_snapshot.replace(" ", "").replace("\u00a0", "")
    if normalized_text in {"계속됩니다", "계속됩니다.", "계속됩니다..", "계속됩니다..."}:
        log_content_debug(product_code, "'계속됩니다' placeholder detected (no other content), skipping.")
        return None

    css_link = soup.new_tag(
        "link",
        rel="stylesheet",
        href="https://static-resource-smartstore.pstatic.net/smartstore/p/static/20230630180923/common.css",
    )
    if soup.head:
        soup.head.append(css_link)
    else:
        head_tag = soup.new_tag("head")
        head_tag.append(css_link)
        soup.insert(0, head_tag)

    for button in soup.find_all("button"):
        button.decompose()

    for img in soup.find_all("img", attrs={"data-src": True}):
        img["src"] = img["data-src"]
        del img["data-src"]

    for img in soup.find_all("img", src="https://rapid-up.s3.ap-northeast-2.amazonaws.com/dev/gray-line.png"):
        img["src"] = _GRAY_LINE_REPLACEMENTS["https://rapid-up.s3.ap-northeast-2.amazonaws.com/dev/gray-line.png"]

    text_to_remove = "* {text-align: center;}  #mycontents11 img{max-width: 100%;}"
    if text_to_remove in soup.get_text():
        soup = BeautifulSoup(str(soup).replace(text_to_remove, ""), "html.parser")
        log_content_debug(product_code, "Removed inline text-align styles.")

    disallowed_attrs = ["area-hidden", "data-linkdata", "data-linktype", "onclick", "style", "class"]
    for attr in disallowed_attrs:
        for tag in soup.find_all(attrs={attr: True}):
            del tag[attr]

    for anchor in soup.find_all("a", attrs={"data-linkdata": True}):
        img = anchor.find("img")
        if not img:
            continue
        src = img.get("src", "")
        data_src = img.get("data-src", "")
        if not src:
            src = data_src
            img["src"] = src
        if "data-src" in img.attrs:
            del img["data-src"]

    for img in soup.find_all("img"):
        src = img.get("src", "")
        if any(src.startswith(prefix) for prefix in _BLOCKED_IMAGE_PREFIXES):
            img.decompose()

    strip_blob_media(soup, product_code)
    cleanup_dom_structure(soup, product_code)

    remaining_img_count = len(soup.find_all('img'))
    log_content_debug(product_code, f"Images after cleanup: {remaining_img_count}")

    soup = insert_and_remove_images(soup)

    for img_tag in soup.find_all("img"):
        img_tag["style"] = "display: block; margin-left: auto; margin-right: auto; margin-bottom: 10px;"

    for h1 in soup.find_all("h1"):
        h1["style"] = "text-align: center; font-size: 30px; margin-bottom: 20px;"

    text_elements = ["p", "div", "span", "li", "a"]
    for tag_name in text_elements:
        for node in soup.find_all(tag_name):
            existing_style = node.get("style", "")
            new_style = f"{existing_style}; text-align: center; font-size: 18px; margin-bottom: 30px;"
            node["style"] = new_style.strip()

    cleaned_html = str(soup).strip()

    meaningful_imgs = [
        img for img in soup.find_all("img")
        if (img.get("src") or "").strip() and (img.get("src").strip() not in BRANDING_IMAGE_URLS)
    ]
    has_text = bool(soup.get_text(strip=True))
    final_html = cleaned_html
    final_label = "cleaned"

    if not has_text and not meaningful_imgs:
        log_content_debug(product_code, "Content empty after cleanup; applying fallback gallery extraction.")
        fallback = build_image_gallery(raw_content, product_code)
        if fallback is None:
            log_content_debug(product_code, "Fallback gallery extraction failed; returning None.")
            save_debug_html(product_code, raw_content, "fallback_failed")
            return None
        final_html = fallback
        final_label = "fallback_gallery"
        log_content_debug(product_code, "Fallback gallery extraction succeeded.")

    if WRAP_CONTENT_HTML:
        final_html = wrap_html_document(final_html)

    dump_content_html(product_code, final_html, final_label)
    log_content_debug(product_code, "Returning cleaned HTML content.")
    return pd.DataFrame({"Content": [final_html]})


def insert_and_remove_images(soup):
    img_srcs_to_insert = [
        "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/top.png",
        "https://axh2eqadoldy.compat.objectstorage.ap-chuncheon-1.oraclecloud.com/bucket-20230610-0005/upload/bottom.png",
        "https://coudae.s3.ap-northeast-2.amazonaws.com/A00412936/cloud/7290.png",
    ]

    img_tag_top = soup.new_tag(
        "img", src=img_srcs_to_insert[0], style="display: block; margin-left: auto; margin-right: auto;"
    )
    img_tag_middle = soup.new_tag(
        "img", src=img_srcs_to_insert[2], style="display: block; margin-left: auto; margin-right: auto;"
    )
    img_tag_bottom = soup.new_tag(
        "img", src=img_srcs_to_insert[1], style="display: block; margin-left: auto; margin-right: auto;"
    )

    try:
        first_tag = next(soup.children)
        last_tag = next(reversed(soup.contents))
    except StopIteration:
        return soup

    first_tag.insert_before(img_tag_top)
    last_tag.insert_after(img_tag_bottom)

    img_srcs_to_remove = ["", ""]
    for img_src in img_srcs_to_remove:
        for img in soup.find_all("img", attrs={"src": img_src}):
            img.decompose()

    return soup


def build_image_gallery(raw_html, product_code="UNKNOWN"):
    if not raw_html:
        log_content_debug(product_code, "build_image_gallery received empty raw_html.")
        return None

    soup = BeautifulSoup(raw_html, 'html.parser')

    for img in soup.find_all('img', attrs={'data-src': True}):
        img['src'] = img['data-src']
        del img['data-src']

    filtered = BeautifulSoup('', 'html.parser')
    container = filtered.new_tag('div')
    filtered.append(container)

    seen = set()
    for img in soup.find_all('img'):
        src = (img.get('src') or '').strip()
        if not src:
            continue
        src = _GRAY_LINE_REPLACEMENTS.get(src, src)
        if any(src.startswith(prefix) for prefix in _BLOCKED_IMAGE_PREFIXES):
            continue
        if src in BRANDING_IMAGE_URLS:
            continue
        if src in seen:
            continue
        seen.add(src)
        clean_img = filtered.new_tag('img', src=src)
        clean_img['style'] = 'display:block;margin:0 auto 10px auto;'
        container.append(clean_img)

    if not container.find_all('img'):
        log_content_debug(product_code, "Fallback gallery extraction produced no images.")
        return None

    log_content_debug(product_code, "Fallback gallery extraction produced image-only content.")
    return str(filtered)


def return_shipping_fee(total_price):
    fee = total_price * 0.25
    if fee > 200000:
        fee = 200000
    return fee


def title_edit(title):
    title_split = title.split(' ')
    title_split = list(dict.fromkeys(title_split))
    if len(title_split) >= 2:
        title_split[-1], title_split[-2] = title_split[-2], title_split[-1]
    title = ' '.join(title_split)
    return title


def get_product_data(page, product, i, num_products):
    title, price, product_url, product_code = extract_product_details(product)

    title = title.replace('\xa0', ' ')

    print(f"Product {i + 1}/{num_products}: {title}, {price} won, {product_url}")

    product_page = context.new_page()
    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(product_page)
    product_page.goto(product_url)
    product_page.wait_for_load_state("load")
    try:
        product_page.wait_for_load_state("networkidle", timeout=10000)
    except PlaywrightTimeoutError:
        pass

    scripts = product_page.query_selector_all('script')

    if not NAVER_CATEGORY_PATH.exists():
        raise FileNotFoundError(f"카테고리 파일을 찾을 수 없습니다: {NAVER_CATEGORY_PATH}")
    category_df = pd.read_excel(NAVER_CATEGORY_PATH, header=None)
    small_category_dict = pd.Series(category_df[0].values, index=category_df[3]).to_dict()
    tiny_category_dict = pd.Series(category_df[0].values, index=category_df[4]).to_dict()

    category = None
    for script in scripts:
        script_content = script.inner_text()
        if "category" in script_content:
            try:
                json_data = json.loads(script_content)
            except Exception:
                continue
            if 'category' in json_data:
                category = json_data['category']
                break

    if category is not None:
        print(f"Category: {category}")
        category_list = category.split(">")
        large_category = category_list[0].strip()
        medium_category = category_list[1].strip()
        small_category = category_list[2].strip() if len(category_list) > 2 else None
        tiny_category = category_list[3].strip() if len(category_list) > 3 else None
    else:
        category_list = []
        large_category = None
        medium_category = None
        small_category = None
        tiny_category = None

    if tiny_category is not None:
        smallest_category = tiny_category
        smallest_category_type = 'tiny'
        naver_category_number = tiny_category_dict.get(tiny_category)
    elif small_category is not None:
        smallest_category = small_category
        smallest_category_type = 'small'
        naver_category_number = small_category_dict.get(small_category)
    else:
        smallest_category = None
        smallest_category_type = None
        naver_category_number = None

    print(
        f"Smallest Category('{smallest_category_type}') : {smallest_category}, Naver category number: {naver_category_number}")

    options = option_crawl(product_page)
    print("Options:", options)

    if not has_numeric_chars(price):
        state_price = price_from_preloaded_state(product_page)
        option_price = price_from_option_data(options)
        resolved_price = None
        source = None
        if state_price:
            resolved_price = state_price
            source = "preloaded_state"
        elif option_price:
            resolved_price = option_price
            source = "option_list"

        if resolved_price:
            price = f"{resolved_price:,}"
            print(f"Price fallback via {source}: {price}")
        else:
            print(f"가격 정보를 찾지 못해 상품을 건너뜁니다: {product_url}")
            product_page.close()
            return None

    common_urls, different_urls = image_crawl(product_page)
    print(f"common_urls: {common_urls}")

    main_image = None
    other_images = []

    if common_urls:
        main_image = common_urls[0].replace('?type=m510', '')
        other_images = [url.replace('?type=m510', '') for url in common_urls[1:]]
        print(f"other_images: {other_images}")
    else:
        print("No common images found")
        try:
            image_element = page.wait_for_selector('xpath=//*[@id="content"]/div/div[2]/div[1]/div[1]/div[1]/img', timeout=2000)
            image_url = image_element.get_attribute("src")
            main_image = image_url.replace('?type=m510', '')
        except Exception:
            print("No main image found")

    print("Main image:", main_image)
    print("Other images:", other_images)

    print("URLs not starting with most common three digits:")
    for url in different_urls:
        print(url)

    element_selector = find_content_element(product_page, product_code)
    content = content_crawl(product_page, product_code, element_selector)
    shipping_fee = original_shipping_fee(product_page)

    def to_int(value):
        if value in (None, "N/A"):
            return 0
        digits = re.sub(r"[^\d]", "", str(value))
        return int(digits) if digits else 0

    shipping_fee_int = to_int(shipping_fee)
    price_int = to_int(price)
    total_price = price_int + shipping_fee_int

    if content is None:
        log_content_debug(product_code, "content_crawl returned None; storing empty placeholder.")
        content_df = pd.DataFrame({'Content': [""]})
    elif isinstance(content, str):
        content_df = pd.DataFrame({'Content': [content]})
    else:
        content_df = content

    product_df = pd.DataFrame({
        'Product': [title],
        'Price': [price],
        'Shipping_Fee': [shipping_fee_int],
        'Total_Price': [total_price],
        'Main_Image': [main_image],
        'Other_Images': [other_images],
        'Options': [options],
        'Product_URL': [product_url],
        'Naver_Category_Number': [naver_category_number]
    })

    product_df = pd.concat([product_df, content_df], axis=1)
    product_page.close()

    return product_df


def write_to_excel(df, excel_path, seen_urls):
    book = load_workbook(excel_path)
    sheet = book['일괄등록']

    b_start_row = c_start_row = e_start_row = h_start_row = ad_start_row = r_start_row = s_start_row = t_start_row = i_start_row = 3
    ap_start_row = aq_start_row = 3

    # 데이터가 없으면 템플릿만 저장하고 조용히 반환
    if df is None or len(df) == 0:
        print("DataFrame이 비어 있어 엑셀 기록을 생략합니다.")
        book.save(excel_path)
        if os.name != "nt":
            print(f"Excel file saved to {excel_path}. (empty dataset)")
        return

    for i, item in enumerate(df['Naver_Category_Number'], start=b_start_row):
        sheet['B' + str(i)] = item
    for j, item in enumerate(df['Product'], start=c_start_row):
        sheet['C' + str(j)] = item
    for k, (product_price, shipping_fee) in enumerate(zip(df['Price'], df['Shipping_Fee']), start=e_start_row):
        total_price = float(product_price.replace(',', '')) + shipping_fee
        selling_price = total_price - 0.01 * total_price
        selling_price_rounded = round(selling_price / 100.0) * 100.0
        sheet['E' + str(k)] = selling_price_rounded

    for l, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['H' + str(l)] = "조합형"
    for m in range(ad_start_row, ad_start_row + len(df)):
        selling_price_rounded = sheet['E' + str(m)].value

        # 바젤마켓 분기
        if selling_price_rounded <= 20000:
            ad_value = 2903608
        elif 20001 <= selling_price_rounded <= 30000:
            ad_value = 2904260
        elif 30001 <= selling_price_rounded <= 40000:
            ad_value = 2904261
        elif 40001 <= selling_price_rounded <= 60000:
            ad_value = 2904262
        elif 60001 <= selling_price_rounded <= 80000:
            ad_value = 2904268
        elif 80001 <= selling_price_rounded <= 100000:
            ad_value = 2904272
        elif 100001 <= selling_price_rounded <= 150000:
            ad_value = 2904276
        elif 150001 <= selling_price_rounded <= 400000:
            ad_value = 2904278
        elif 400001 <= selling_price_rounded <= 600000:
            ad_value = 2904279
        elif 600001 <= selling_price_rounded <= 1000000:
            ad_value = 2904281
        elif 1000001 <= selling_price_rounded <= 9999999:
            ad_value = 2904284

        sheet['AD' + str(m)] = ad_value

    for row, _ in enumerate(df.iterrows(), start=h_start_row):
        sheet['U' + str(row)] = "상세페이지 참조"
        sheet['V' + str(row)] = "상세페이지 참조"
        sheet['Y' + str(row)] = "0200037"
        sheet['Z' + str(row)] = "구매대행"
        sheet['AZ' + str(row)] = "010-3973-3119"
        sheet['BA' + str(row)] = "본문 안내문 참조"

    for l, item in enumerate(df['Options'], start=r_start_row):
        option_titles = []
        option_prices = []
        option_categories = []
        for key in item:
            option_categories.append(key)
            if '하위옵션제목' in item[key]:
                option_titles.append(', '.join(item[key]['하위옵션제목']))
            if '하위옵션가격' in item[key]:
                option_prices.append(', '.join(map(str, item[key]['하위옵션가격'])))
        sheet['I' + str(l)] = '\n'.join(option_categories)
        sheet['J' + str(l)] = '\n'.join(option_titles)
        sheet['K' + str(l)] = '\n'.join(option_prices)

    if 'Options' in df.columns:
        for row in range(h_start_row, h_start_row + len(df)):
            item_options = df.at[row - h_start_row, 'Options']
            option_prices = []
            for option in item_options.values():
                if '하위옵션가격' in option:
                    option_prices.extend(option['하위옵션가격'])

            if option_prices:
                num_prices = len(option_prices)
                l_values = ', '.join(['99'] * num_prices)
                sheet['L' + str(row)] = l_values
            else:
                sheet['L' + str(row)] = "99"
    else:
        for row in range(h_start_row, h_start_row + len(df)):
            sheet['L' + str(row)] = "99"

    for n, item in enumerate(df['Main_Image'], start=r_start_row):
        sheet['R' + str(n)] = item
    for o, item in enumerate(df['Other_Images'], start=s_start_row):
        if item is not None:
            sheet['S' + str(o)] = "\n".join(str(img) for img in item if img is not None)
        else:
            sheet['S' + str(o)] = ""

    if 'Content' in df.columns:
        for p, item in enumerate(df['Content'], start=t_start_row):
            if isinstance(item, float):
                item = str(item)
            print(f"Row {p}, Content: {item[:100]}")
            sheet['T' + str(p)] = item
    else:
        print("No 'Content' column found in DataFrame")

    for q, url in enumerate(df['Product_URL'], start=i_start_row):
        product_code = url.split('/')[-1]
        try:
            selling_code = str(int(product_code) * 2)
        except ValueError:
            selling_code = str(random.randint(10000000, 99999999)) + 'R'
        sheet['A' + str(q)] = selling_code
    for r, total_price in enumerate(df['Total_Price'], start=ap_start_row):
        return_fee = return_shipping_fee(total_price)
        return_fee_rounded = round(return_fee / 100.0) * 100
        sheet['AP' + str(r)] = return_fee_rounded
        sheet['AQ' + str(r)] = return_fee_rounded * 2

    book.save(excel_path)

    # 뒤쪽 빈 행 제거
    book = load_workbook(excel_path)
    sheet = book['일괄등록']
    for i in range(3, sheet.max_row + 1):
        if not sheet['A' + str(i)].value:
            sheet.delete_rows(i, sheet.max_row - i + 1)
            break
    book.save(excel_path)

    if os.name == "nt":
        os.system(f'start "" "excel.exe" "{excel_path}"')
    else:
        print(f"Excel file saved to {excel_path}. Automatic Excel launch is skipped on non-Windows platforms.")


def write_to_excel2(df, excel_path2):
    df2 = pd.DataFrame({
        'Product_URL': df['Product_URL'],
        'Numbering': range(1, len(df) + 1),
        'Product_Title': df['Product'],
        'Product_Price': df['Price'],
        'Shipping_Fee': df['Shipping_Fee']
    })
    with pd.ExcelWriter(excel_path2) as writer:
        df2.to_excel(writer, index=False)


# 실행부: 항상 로컬 브라우저를 실행 (Windows 우선)
if CRAWLER_DRY_RUN:
    print("CRAWLER_DRY_RUN=1 플래그로 인해 Playwright 크롤링 본동작을 생략합니다.")
    sys.stdout = original
    f.close()
    sys.exit(0)

with sync_playwright() as p:
    browser_name = os.getenv("PLAYWRIGHT_BROWSER", "chromium").lower()
    if browser_name not in {"chromium", "firefox", "webkit"}:
        browser_name = "chromium"

    headless_mode = os.getenv("PLAYWRIGHT_HEADLESS", "0").lower() in {"1", "true", "yes"}

    if browser_name == "chromium":
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-features=NetworkService",
            "--disable-web-security",
            "--disable-dev-shm-usage",
            "--disable-accelerated-2d-canvas",
            "--disable-gpu",
        ]
        browser = p.chromium.launch(headless=headless_mode, args=launch_args)
        context = browser.new_context()
    elif browser_name == "firefox":
        browser = p.firefox.launch(headless=headless_mode)
        context = browser.new_context()
    else:
        browser = p.webkit.launch(headless=headless_mode)
        context = browser.new_context()

    if browser_name == "chromium" and STEALTH_HELPER:
        STEALTH_HELPER.apply_stealth_sync(context)

    df_columns = [
        'Naver_Category_Number',
        'Product',
        'Price',
        'Shipping_Fee',
        'Total_Price',
        'Options',
        'Main_Image',
        'Other_Images',
        'Content',
        'Product_URL',
    ]
    df = pd.DataFrame(columns=df_columns)
    output_folder = SCRIPT_DIR / 'output'
    read_excel_path = output_folder / 'ExcelSaveTemplate_230109.xlsx'
    seen_urls = set()

    product_list_crawl(context, df, read_excel_path, seen_urls)
    try:
        context.close()
    finally:
        browser.close()

sys.stdout = original
f.close()