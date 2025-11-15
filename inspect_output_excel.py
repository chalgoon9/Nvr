from pathlib import Path

from openpyxl import load_workbook


def main():
    base = Path("output")
    all_candidates = sorted(base.glob("dolce_joypapa__ALL_61_61_*.xlsx"))
    # _second 가 붙지 않은 것이 메인 결과 파일
    main_candidates = [p for p in all_candidates if not p.name.endswith("_second.xlsx")]
    second_candidates = [p for p in all_candidates if p.name.endswith("_second.xlsx")]
    if not main_candidates:
        print("No main dolce_joypapa__ALL_61_61_*.xlsx file found.")
        return
    main_path = main_candidates[-1]
    second_path = second_candidates[-1] if second_candidates else None

    print("MAIN", main_path)
    print("SECOND", second_path)

    wb = load_workbook(main_path)
    print("SHEETS", wb.sheetnames)
    # 대부분 '요약시트'지만, 다른 이름일 수도 있어 첫 번째 시트를 기본으로 사용
    sheet_name = "요약시트"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    print("SUMMARY_ROWS")
    for row in range(3, 20):
        a = ws[f"A{row}"].value
        c = ws[f"C{row}"].value
        t = ws[f"T{row}"].value
        if a is None and c is None and t is None:
            continue
        if isinstance(t, str) and len(t) > 80:
            t_snip = t[:80] + "..."
        else:
            t_snip = t
        print(row, repr(a), repr(c), repr(t_snip))

    print("VIDEO_ROWS")
    for row in range(3, 20):
        t = ws[f"T{row}"].value or ""
        if isinstance(t, str) and ("se-component se-video" in t or "<video" in t):
            print(row, repr(t[:120]))

    # 문제 되었던 플레이어 UI 텍스트가 남아있는지 확인
    problem_keywords = ["글자 크기", "배경색", "라이선스", "사용 안함", "0초", "720p", "480p", "270p", "HD", "1.0x"]
    print("PROBLEM_KEYWORDS_IN_CONTENT")
    for row in range(3, 20):
        t = ws[f"T{row}"].value or ""
        if not isinstance(t, str):
            continue
        hits = [k for k in problem_keywords if k in t]
        if hits:
            print("ROW", row, "HITS", hits)

    if second_path:
        wb2 = load_workbook(second_path)
        ws2 = wb2.active
        print("SECOND_ROWS")
        for row in range(2, 10):
            vals = [ws2[f"{col}{row}"].value for col in ["A", "B", "C", "D", "E"]]
            if any(vals):
                print(row, repr(vals))


if __name__ == "__main__":
    main()
